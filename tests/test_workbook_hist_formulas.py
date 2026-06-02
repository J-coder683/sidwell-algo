"""
test_workbook_hist_formulas.py — Part A regression: historical ratio cells in the
Income Statement and Balance Sheet must be live Excel FORMULAS, not static values.

openpyxl reads cell content — if the value starts with '=' the cell is a formula.
We verify that the builder wrote formula strings (not Python-computed numbers) for
the derived/ratio historical cells, mirroring the projection-column pattern.
"""

import io
import pytest
import openpyxl

from sidwell.render.workbook import WorkbookRenderer
from sidwell.engine.statements import StatementsEngine
from valuation.dcf import run_dcf_valuation
from sidwell.ajp.schema import AJP, AJPMeta

_ERROR_TOKENS = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!", "#N/A")


def _make_ajp():
    return AJP(
        meta=AJPMeta(
            ticker="HISTFML", as_of="2025-03-31", currency="INR_MM",
            sources_ingested=[], fiscal_year_end_month=3,
            last_actual_fy="FY2025", is_holdco=False, scenario_active="BASE",
        ),
        assumptions=[],
    )


def _rich_financials():
    """Fixture with realistic non-zero historical data (4 years) so all formula paths fire."""
    return {
        "ticker": "HISTFML",
        "current_price": 100.0,
        "market_cap": 5000.0,
        "shares_outstanding": 50.0,
        "years": ["2022", "2023", "2024", "2025"],
        "revenue": [1000.0, 1100.0, 1210.0, 1331.0],
        "gross_profit": [400.0, 440.0, 484.0, 532.4],
        "ebit": [150.0, 165.0, 181.5, 199.65],
        "interest_expense": [0.0] * 4,
        "tax_provision": [30.0, 33.0, 36.3, 39.93],
        "pretax_income": [150.0, 165.0, 181.5, 199.65],
        "net_income": [120.0, 132.0, 145.2, 159.72],
        "total_assets": [800.0] * 4,
        "total_equity": [600.0] * 4,
        "cash": [100.0] * 4,
        "debt": [50.0] * 4,
        "capex": [40.0, 44.0, 48.4, 53.24],
        "depreciation": [20.0, 22.0, 24.2, 26.62],
        "working_capital_change": [0.0] * 4,
        "fcf": [100.0] * 4,
        "statements": {
            "years_annual": ["Mar 2022", "Mar 2023", "Mar 2024", "Mar 2025"],
            "annual": {
                "profit_loss": {
                    "sales":             [100.0, 110.0, 121.0, 133.1],
                    "cogs":              [60.0, 66.0, 72.6, 79.86],
                    "operating profit":  [15.0, 16.5, 18.15, 19.965],
                    "depreciation":      [2.0, 2.2, 2.42, 2.662],
                    "interest":          [0.0, 0.0, 0.0, 0.0],
                    "profit before tax": [15.0, 16.5, 18.15, 19.965],
                    "tax":               [3.0, 3.3, 3.63, 3.993],
                    "tax %":             [20.0, 20.0, 20.0, 20.0],
                    "net profit":        [12.0, 13.2, 14.52, 15.972],
                    "dividend payout %": [40.0, 42.0, 45.0, 44.0],
                },
                "balance_sheet": {
                    "equity capital":    [50.0, 52.0, 54.0, 56.0],
                    "reserves":          [10.0, 11.0, 12.0, 13.0],
                    "borrowings":        [5.0, 5.0, 5.0, 5.0],
                    "fixed assets":      [30.0, 32.0, 34.0, 36.0],
                    "trade payables":    [8.0, 8.8, 9.68, 10.648],
                    "inventories":       [10.0, 11.0, 12.1, 13.31],
                    "trade receivables": [12.0, 13.2, 14.52, 15.972],
                    "cash equivalents":  [5.0, 5.5, 6.05, 6.655],
                },
                "cash_flow": {
                    "fixed assets purchased": [3.0, 3.3, 3.63, 3.993],
                    "cash from operating activity": [13.0, 14.3, 15.73, 17.303],
                },
            },
            "ratios": {
                "debtor days":    [43.0, 43.0, 43.0, 43.0],
                "inventory days": [60.0, 60.0, 60.0, 60.0],
                "days payable":   [48.0, 48.0, 48.0, 48.0],
            },
        },
    }


@pytest.fixture(scope="module")
def wb_rich():
    fin = _rich_financials()
    dcf_res = run_dcf_valuation(fin, {}, 0.04, None)
    engine_results = dcf_res["engine_results"]
    renderer = WorkbookRenderer(engine_results, _make_ajp())
    wb = renderer.render()
    buf = io.BytesIO()
    wb.save(buf)
    return openpyxl.load_workbook(io.BytesIO(buf.getvalue()))


# ── IS historical formula cells ───────────────────────────────────────────────

N_HIST = 4   # 4 historical years in the fixture

# Row constants matching render_is layout
R_G, R_REV, R_COGSP, R_COGS = 3, 4, 5, 6
R_MG, R_EBIT, R_TAX, R_NOP  = 7, 8, 9, 10
R_CXP, R_CX, R_DAP, R_DA, R_NP = 11, 12, 13, 14, 15


def _is_ws(wb_rich):
    return wb_rich["4_Income_Statement"]


def test_is_hist_revenue_growth_is_formula(wb_rich):
    """Historical Revenue growth % cells (from col 4 onward) must be formulas."""
    ws = _is_ws(wb_rich)
    # First col (col 3) is blank (no prior year); cols 4..3+N_HIST-1 should be formulas
    formula_count = sum(
        1 for i in range(1, N_HIST)   # skip i=0 (no prior year)
        if str(ws.cell(row=R_G, column=3 + i).value or "").startswith("=")
    )
    assert formula_count >= N_HIST - 1, (
        f"Expected {N_HIST-1} revenue growth formulas, got {formula_count}"
    )


def test_is_hist_cogs_pct_is_formula(wb_rich):
    ws = _is_ws(wb_rich)
    assert any(
        str(ws.cell(row=R_COGSP, column=3 + i).value or "").startswith("=")
        for i in range(N_HIST)
    ), "Expected COGS % historical cells to be formulas"


def test_is_hist_ebit_margin_is_formula(wb_rich):
    ws = _is_ws(wb_rich)
    assert any(
        str(ws.cell(row=R_MG, column=3 + i).value or "").startswith("=")
        for i in range(N_HIST)
    ), "Expected EBIT margin % historical cells to be formulas"


def test_is_hist_capex_pct_is_formula(wb_rich):
    ws = _is_ws(wb_rich)
    assert any(
        str(ws.cell(row=R_CXP, column=3 + i).value or "").startswith("=")
        for i in range(N_HIST)
    ), "Expected CapEx % historical cells to be formulas"


def test_is_hist_nopat_is_formula(wb_rich):
    ws = _is_ws(wb_rich)
    nopat_formulas = [
        ws.cell(row=R_NOP, column=3 + i).value
        for i in range(N_HIST)
    ]
    assert any(str(v or "").startswith("=") for v in nopat_formulas), (
        f"NOPAT historical cells not formulas: {nopat_formulas}"
    )
    for v in nopat_formulas:
        if str(v or "").startswith("="):
            for tok in _ERROR_TOKENS:
                assert tok not in str(v), f"Error token in NOPAT formula: {v}"


def test_is_hist_da_rate_is_formula(wb_rich):
    """D&A rate on PP&E: from 2nd hist column onward must be formulas."""
    ws = _is_ws(wb_rich)
    formula_count = sum(
        1 for i in range(1, N_HIST)   # first col has no prior PPE
        if str(ws.cell(row=R_DAP, column=3 + i).value or "").startswith("=")
    )
    assert formula_count >= N_HIST - 2, (
        f"Expected D&A rate historical formulas from col 2 onward, got {formula_count}"
    )


def test_no_ref_errors_in_is_hist(wb_rich):
    """No #REF! / #NAME? tokens in any IS cell's formula string."""
    ws = _is_ws(wb_rich)
    bad = [
        f"{cell.coordinate}: {cell.value}"
        for row in ws.iter_rows()
        for cell in row
        if str(cell.value or "").startswith("=")
        for tok in ("#REF!", "#NAME?")
        if tok in str(cell.value)
    ]
    assert not bad, f"Error tokens in IS formulas: {bad}"


# ── BS historical formula cells ───────────────────────────────────────────────

# Row constants matching new render_bs layout (with dividend rows inserted)
R_DSO, R_DIO, R_DPO = 3, 4, 5
R_CASH_BS, R_AR, R_INV, R_OWC, R_PPE, R_OTH, R_TA = 6, 7, 8, 9, 10, 11, 12
R_AP, R_DEBT_BS, R_EQ = 13, 14, 15
R_PAYP, R_DIV = 16, 17
R_TLE, R_NWC, R_BC = 18, 19, 20


def _bs_ws(wb_rich):
    return wb_rich["5_Balance_Sheet"]


def test_bs_hist_total_assets_is_formula(wb_rich):
    ws = _bs_ws(wb_rich)
    assert any(
        str(ws.cell(row=R_TA, column=3 + i).value or "").startswith("=")
        for i in range(N_HIST)
    ), "Expected Total Assets historical cells to be formulas"


def test_bs_hist_total_le_is_formula(wb_rich):
    ws = _bs_ws(wb_rich)
    assert any(
        str(ws.cell(row=R_TLE, column=3 + i).value or "").startswith("=")
        for i in range(N_HIST)
    ), "Expected Total L+E historical cells to be formulas"


def test_bs_hist_balance_check_is_formula(wb_rich):
    ws = _bs_ws(wb_rich)
    assert any(
        str(ws.cell(row=R_BC, column=3 + i).value or "").startswith("=")
        for i in range(N_HIST)
    ), "Expected Balance Check historical cells to be formulas"


def test_bs_hist_dso_is_formula(wb_rich):
    """DSO historical cells must be formula when AR and Revenue are non-zero."""
    ws = _bs_ws(wb_rich)
    assert any(
        str(ws.cell(row=R_DSO, column=3 + i).value or "").startswith("=")
        for i in range(N_HIST)
    ), "Expected DSO historical cells to be formulas"


def test_bs_hist_nwc_is_formula(wb_rich):
    ws = _bs_ws(wb_rich)
    assert any(
        str(ws.cell(row=R_NWC, column=3 + i).value or "").startswith("=")
        for i in range(N_HIST)
    ), "Expected NWC historical cells to be formulas"


def test_no_ref_errors_in_bs_hist(wb_rich):
    """No #REF! / #NAME? tokens in any BS cell's formula string."""
    ws = _bs_ws(wb_rich)
    bad = [
        f"{cell.coordinate}: {cell.value}"
        for row in ws.iter_rows()
        for cell in row
        if str(cell.value or "").startswith("=")
        for tok in ("#REF!", "#NAME?")
        if tok in str(cell.value)
    ]
    assert not bad, f"Error tokens in BS formulas: {bad}"


def test_bs_proj_equity_roll_contains_payout(wb_rich):
    """Projected Equity roll formula must reference (1-payout)."""
    ws = _bs_ws(wb_rich)
    proj_col = 3 + N_HIST   # first projection column
    val = str(ws.cell(row=R_EQ, column=proj_col).value or "")
    assert "1-" in val or "(1" in val, (
        f"Equity proj formula doesn't subtract payout: {val!r}"
    )


def test_bs_proj_cash_roll_contains_dividend(wb_rich):
    """Projected Cash roll formula must reference the Dividends row (R_DIV=17)."""
    ws = _bs_ws(wb_rich)
    proj_col = 3 + N_HIST
    val = str(ws.cell(row=R_CASH_BS, column=proj_col).value or "")
    assert str(R_DIV) in val, (
        f"Cash proj formula doesn't reference Dividends row {R_DIV}: {val!r}"
    )


def test_bs_div_payout_row_exists(wb_rich):
    """Row 16 label must be 'Dividend payout %'."""
    ws = _bs_ws(wb_rich)
    label = ws.cell(row=R_PAYP, column=2).value
    assert label and "payout" in str(label).lower(), (
        f"Expected 'Dividend payout %' label at row {R_PAYP}, got {label!r}"
    )


def test_bs_dividends_row_exists(wb_rich):
    """Row 17 label must be 'Dividends'."""
    ws = _bs_ws(wb_rich)
    label = ws.cell(row=R_DIV, column=2).value
    assert label and "dividend" in str(label).lower(), (
        f"Expected 'Dividends' label at row {R_DIV}, got {label!r}"
    )
