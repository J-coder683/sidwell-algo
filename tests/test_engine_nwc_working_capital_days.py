"""
test_engine_nwc_working_capital_days.py — NWC driven by screener's signed
'Working Capital Days', using DSO/DIO/DPO for the itemized trade legs.

Why: screener publishes Debtor Days but frequently leaves Inventory Days and
Days Payable blank (telecom, services). The engine used to fabricate those
(30/45 days) and reconstruct NWC = AR + Inv − AP. That (a) invents working
capital that doesn't exist and (b) misses the large "other" current items that
make capital-light businesses run *negative* working capital.

Evidence (cached screener ratios, last FY):
    TATACOMM: Debtor 61, Inv/Payable BLANK, Working Capital Days −116
    RELIANCE: Debtor 20, Inv 88, Payable 84  → CCC +24, but WC Days −32
    ASIANPAINT: Debtor 46, Inv 105, Payable 70 → CCC +81, but WC Days +102
The three trade-day lines never reconcile to Working Capital Days — the gap is
real "other net working capital" (advances, deferred revenue, other liabs).

Model under test (non-financial): AR/Inv/AP from their *real* scraped days
(blank → 0, never fabricated), and an "other net WC" line so the comprehensive
net equals Working Capital Days × sales. ΔNWC (the cash-flow term) uses that
comprehensive net, anchored on the same basis so there is no Year-1 jump.
"""

import pytest
from sidwell.ajp.schema import AJP, AJPMeta
from sidwell.engine.statements import StatementsEngine


def _meta() -> AJPMeta:
    return AJPMeta(
        ticker="TATACOMM", as_of="2026-03-31", currency="INR_MM",
        sources_ingested=[], fiscal_year_end_month=3,
        last_actual_fy="FY2026", is_holdco=False, scenario_active="BASE",
    )


def _ajp_blank() -> AJP:
    return AJP(meta=_meta(), assumptions=[])


def _tatacomm_statements(with_wc_days=True, is_financial=False):
    """Real TATACOMM actuals (crore → mm ×10). Telecom: negative working capital,
    blank Inventory/Payable days, populated Working Capital Days."""
    return {
        "years_annual": ["Mar 2022", "Mar 2023", "Mar 2024", "Mar 2025", "Mar 2026"],
        "annual": {
            "profit_loss": {
                "sales":             [16725.0, 17838.0, 20969.0, 23109.0, 24803.0],
                "operating profit":  [4227.0, 4318.0, 4230.0, 4569.0, 4822.0],
                "depreciation":      [2205.0, 2262.0, 2470.0, 2592.0, 2827.0],
                "interest":          [0.0, 0.0, 0.0, 0.0, 0.0],
                "profit before tax": [2000.0, 2063.0, 1163.0, 2281.0, 1396.0],
                "tax":               [515.0, 262.0, 193.0, 444.0, 399.0],
                "net profit":        [1485.0, 1801.0, 970.0, 1837.0, 997.0],
            },
            "balance_sheet": {
                "equity capital":   [285.0, 285.0, 285.0, 285.0, 285.0],
                "reserves":         [643.0, 1233.0, 1501.0, 2736.0, 3162.0],
                "borrowings":       [9122.0, 8577.0, 11263.0, 12357.0, 12249.0],
                "fixed assets":     [10936.0, 10500.0, 13467.0, 13919.0, 15570.0],
                "inventories":      [38.0, 160.0, 84.0, 158.0, 100.0],
                "trade receivables":[2582.0, 2735.0, 3758.0, 4006.0, 4171.0],
                "trade payables":   [3006.0, 3277.0, 3656.0, 3569.0, 4317.0],
                "cash equivalents": [743.0, 1063.0, 842.0, 587.0, 695.0],
            },
            "cash_flow": {
                "fixed assets purchased": [-2500.0, -2600.0, -3200.0, -3400.0, -3800.0],
            },
        },
        "ratios": {
            "debtor days":  [56.0, 65.0, 63.0, 61.0, 61.0],
            "inventory days": [None, None, None, None, None],
            "days payable": [None, None, None, None, None],
            "working capital days": (
                [-119.0, -151.0, -107.0, -116.0, -116.0] if with_wc_days else []
            ),
        },
    }


def _tatacomm_hist(**kw):
    return StatementsEngine.map_historical(_tatacomm_statements(**kw))


# ---------------------------------------------------------------------------
def test_working_capital_days_drives_negative_nwc():
    """NWC level tracks signed Working Capital Days (negative for telecom)."""
    hist = _tatacomm_hist()
    proj = StatementsEngine.run_projections(hist, _ajp_blank())

    # Every projected NWC is materially negative (≈ −116/365 × sales).
    for i, nwc in enumerate(proj["nwc"]):
        implied = (-116.0 / 365.0) * proj["revenue"][i]
        assert nwc < 0, f"year {i}: expected negative NWC, got {nwc:.0f}"
        # within 25% of the WC-days implied level (margins/growth aside)
        assert abs(nwc - implied) < 0.25 * abs(implied), (
            f"year {i}: NWC {nwc:.0f} not near WC-days implied {implied:.0f}"
        )


def test_no_year1_nwc_phantom_and_releases_cash():
    """Continuous anchor → no Year-1 spike; negative WC grows → cash source."""
    hist = _tatacomm_hist()
    proj = StatementsEngine.run_projections(hist, _ajp_blank())

    rev1 = proj["revenue"][0]
    # No phantom: Year-1 ΔNWC is a small fraction of revenue and same sign as Y2.
    assert abs(proj["nwc_change"][0]) < 0.10 * rev1
    assert proj["nwc_change"][0] < 0 and proj["nwc_change"][1] < 0

    # Negative, growing WC releases cash: UFCF exceeds NOPAT + D&A − CapEx.
    for i in range(len(proj["ufcf"])):
        plain = proj["nopat"][i] + proj["da"][i] - proj["capex"][i]
        assert proj["ufcf"][i] > plain - 1.0


def test_other_wc_reconciles_to_working_capital_days():
    """AR + Inv − AP + other_wc must equal the comprehensive NWC every year."""
    hist = _tatacomm_hist()
    proj = StatementsEngine.run_projections(hist, _ajp_blank())
    for i in range(len(proj["nwc"])):
        recon = proj["ar"][i] + proj["inv"][i] - proj["ap"][i] + proj["other_wc"][i]
        assert abs(recon - proj["nwc"][i]) < 1.0


def test_balance_sheet_still_ties_with_other_wc():
    hist = _tatacomm_hist()
    proj = StatementsEngine.run_projections(hist, _ajp_blank())
    for i, bc in enumerate(proj["balance_check"]):
        assert abs(bc) < 1.0, f"balance check failed in year {i}: {bc:.4f}mm"


def test_blank_days_are_not_fabricated():
    """With no Working Capital Days and blank Inv/Payable days, the engine must
    use 0 days for the missing lines — never the old 30/45 fabrication."""
    hist = _tatacomm_hist(with_wc_days=False)
    proj = StatementsEngine.run_projections(hist, _ajp_blank())
    au = proj["assumptions_used"]
    assert au["dio_days"] == 0.0, f"inventory days must be 0 when blank, got {au['dio_days']}"
    assert au["dpo_days"] == 0.0, f"payable days must be 0 when blank, got {au['dpo_days']}"


def test_workbook_bs_mirrors_working_capital_days_model():
    """The 13-sheet workbook's live BS must mirror the engine: NWC linked to the
    signed WC-Days row, AP derived (AR+Inv−NWC), and the FCF-sheet ΔNWC equal to
    the engine's (negative for TATACOMM's negative working capital)."""
    from valuation.dcf import run_dcf_valuation
    from sidwell.render.workbook import WorkbookRenderer

    fin = {
        "ticker": "TATACOMM", "current_price": 1700.0, "market_cap": 4_800_000.0,
        "shares_outstanding": 2850.0, "is_bank": False, "is_financial": False,
        "fcf": [1000.0, 1500.0, 1800.0, 2000.0],
        "statements": _tatacomm_statements(),
    }
    res = run_dcf_valuation(fin, {}, 0.07, None)
    wb = WorkbookRenderer(res["engine_results"], res["ajp"]).render()
    bs = wb["5_Balance_Sheet"]

    assert bs.cell(row=5, column=2).value == "DPO (days)"
    assert bs.cell(row=9, column=2).value == "Other net working capital"
    # first projection column
    c = 3
    while not (isinstance(bs.cell(row=2, column=c).value, str)
               and bs.cell(row=2, column=c).value.rstrip().endswith("E")):
        c += 1

    from openpyxl.utils import get_column_letter as L
    col = L(c)
    nwc_formula = str(bs.cell(row=19, column=c).value)   # R_NWC (shifted +2 after dividend rows 16,17)
    owc_formula = str(bs.cell(row=9, column=c).value)    # R_OWC (unchanged)
    ap_formula = str(bs.cell(row=13, column=c).value)    # R_AP (unchanged)
    
    assert "/365" in nwc_formula, f"NWC must use wc_days: {nwc_formula}"
    assert f"{col}19-({col}7+{col}8-{col}13)" in owc_formula, f"OWC must be derived: {owc_formula}"
    assert f"{col}5/365" in ap_formula, f"AP must be independent: {ap_formula}"

    # FCF sheet ΔNWC equals the engine's and is negative (cash released).
    fcf = wb["8_FCF_DCF"]
    fc = 3
    while not (isinstance(fcf.cell(row=2, column=fc).value, str)
               and fcf.cell(row=2, column=fc).value.rstrip().endswith("E")):
        fc += 1
    eng_dnwc = res["engine_results"]["proj"]["nwc_change"][0]
    assert eng_dnwc < 0
    assert abs(fcf.cell(row=4, column=fc).value - eng_dnwc) < 1.0


def test_financial_freeze_overrides_working_capital_days():
    """A financial with populated Working Capital Days is still frozen (ΔNWC=0)."""
    hist = _tatacomm_hist()
    proj = StatementsEngine.run_projections(
        hist, _ajp_blank(), freeze_working_capital=True
    )
    for i, d in enumerate(proj["nwc_change"]):
        assert abs(d) < 1.0, f"frozen ΔNWC must be 0 in year {i}, got {d:.2f}"
