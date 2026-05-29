"""
test_excel_recalc.py — Structural workbook integrity check

IMPORTANT SCOPE LIMITATION:
    This test is a STRUCTURAL CHECK, NOT a formula-correctness gate.
    openpyxl reads formulas as strings and never evaluates them.
    It cannot detect #DIV/0! or #VALUE! errors that only appear after
    Excel/LibreOffice evaluates the workbook.

    The NUMERIC CORRECTNESS gate is test_acceptance.py:TestMathReconciliation,
    which independently recomputes EV/equity/per-share from engine primitives.

What this test DOES catch:
    - #REF! / #NAME? / #VALUE! text literally baked into formula strings
      (e.g., someone accidentally wrote '=#REF!' as a formula)
    - Cells marked as formulas (data_type == 'f') with missing or empty content
    - Sheets with no content at all (structural build failure)

What this test CANNOT catch:
    - Runtime formula evaluation errors (requires Excel or LibreOffice headless)
    - Circular references causing #VALUE! at evaluation time
    - Formulas that produce the wrong number

Upgrade path: replace with a headless LibreOffice recalc step in CI using
    soffice --headless --convert-to xlsx <file.xlsx>
or use the `formulas` / `xlcalculator` library if available.
"""

import io
import pytest
import openpyxl

from valuation.dcf import run_dcf_valuation
from exports.excel import export_dcf_excel


_ERROR_TOKENS = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!", "#N/A")


def _build_workbook() -> openpyxl.Workbook:
    fin = {
        "ticker": "XLTEST",
        "current_price": 10.0,
        "market_cap": 100.0,
        "shares_outstanding": 10.0,
        "years": ["2021", "2022", "2023", "2024"],
        "revenue": [110.0, 121.0, 133.1, 146.41],
        "gross_profit": [55.0, 60.5, 66.55, 73.205],
        "ebit": [16.5, 18.15, 19.965, 21.9615],
        "interest_expense": [0.0, 0.0, 0.0, 0.0],
        "tax_provision": [4.15, 4.57, 5.03, 5.53],
        "pretax_income": [16.5, 18.15, 19.965, 21.9615],
        "net_income": [12.35, 13.58, 14.94, 16.43],
        "total_assets": [100.0, 100.0, 100.0, 100.0],
        "total_equity": [80.0, 80.0, 80.0, 80.0],
        "cash": [20.0, 20.0, 20.0, 20.0],
        "debt": [0.0, 0.0, 0.0, 0.0],
        "capex": [3.3, 3.63, 3.993, 4.3923],
        "depreciation": [2.2, 2.42, 2.662, 2.9282],
        "working_capital_change": [0.0, 0.0, 0.0, 0.0],
        "fcf": [11.25, 12.37, 13.61, 14.97],
        "statements": {
            "years_annual": ["2022", "2023", "2024", "2025"],
            "annual": {
                "profit_loss": {
                    "sales":            [11.0, 12.1, 13.31, 14.641],
                    "operating profit": [1.65, 1.815, 1.9965, 2.19615],
                    "depreciation":     [0.22, 0.242, 0.2662, 0.29282],
                    "interest":         [0.0, 0.0, 0.0, 0.0],
                    "profit before tax":[1.65, 1.815, 1.9965, 2.19615],
                    "tax":              [0.415, 0.456, 0.502, 0.552],
                    "net profit":       [1.23, 1.35, 1.49, 1.64],
                },
                "balance_sheet": {
                    "equity capital":   [8.0, 8.0, 8.0, 8.0],
                    "reserves":         [0.0, 0.0, 0.0, 0.0],
                    "borrowings":       [0.0, 0.0, 0.0, 0.0],
                    "fixed assets":     [5.0, 5.0, 5.0, 5.0],
                    "trade payables":   [0.0, 0.0, 0.0, 0.0],
                    "inventories":      [0.0, 0.0, 0.0, 0.0],
                    "trade receivables":[0.0, 0.0, 0.0, 0.0],
                    "cash equivalents": [2.0, 2.0, 2.0, 2.0],
                },
                "cash_flow": {
                    "fixed assets purchased": [0.33, 0.363, 0.3993, 0.43923],
                },
            },
            "ratios": {},
        },
    }
    dcf_res = run_dcf_valuation(fin, {}, 0.04, None)
    excel_bytes = export_dcf_excel(dcf_res, fin)
    return openpyxl.load_workbook(io.BytesIO(excel_bytes))


@pytest.fixture(scope="module")
def wb():
    return _build_workbook()


# ---------------------------------------------------------------------------
# Structural checks
# ---------------------------------------------------------------------------

def test_workbook_has_multiple_sheets(wb):
    """At least 3 sheets must be present (summary, projections, assumptions)."""
    assert len(wb.sheetnames) >= 3, (
        f"Expected ≥3 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"
    )


def test_no_error_tokens_in_formula_strings(wb):
    """
    STRUCTURAL CHECK (not a correctness gate — see module docstring).
    Scan all formula strings for error tokens.  These would indicate a broken
    formula string was literally written to the file, not a runtime eval error.
    """
    violations = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula_str = str(cell.value)
                    for token in _ERROR_TOKENS:
                        if token in formula_str:
                            violations.append(
                                f"{sheet_name}!{cell.coordinate}: "
                                f"formula contains '{token}'"
                            )

    assert len(violations) == 0, (
        "Formula string error tokens found (structural build bug):\n"
        + "\n".join(violations[:20])
    )


def test_no_empty_formula_cells(wb):
    """
    Cells marked as formula (data_type == 'f') must not be empty strings.
    An empty formula cell indicates a builder bug where = was written without content.
    """
    violations = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    if not cell.value or str(cell.value).strip() in ("=", ""):
                        violations.append(f"{sheet_name}!{cell.coordinate}")

    assert len(violations) == 0, (
        f"Found {len(violations)} empty formula cells: {violations[:10]}"
    )


def test_all_sheets_have_content(wb):
    """Every sheet must have at least one non-None cell."""
    empty_sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        has_content = any(
            cell.value is not None
            for row in ws.iter_rows()
            for cell in row
        )
        if not has_content:
            empty_sheets.append(sheet_name)

    assert len(empty_sheets) == 0, (
        f"These sheets are completely empty: {empty_sheets}"
    )


def test_output_is_valid_xlsx_bytes():
    """export_dcf_excel must return bytes that openpyxl can load."""
    fin = {
        "ticker": "XLTEST2", "current_price": 10.0, "market_cap": 100.0,
        "shares_outstanding": 10.0,
        "revenue": [110.0], "ebit": [16.5], "interest_expense": [0.0],
        "tax_provision": [4.15], "pretax_income": [16.5], "net_income": [12.35],
        "total_assets": [100.0], "total_equity": [80.0], "cash": [20.0],
        "debt": [0.0], "capex": [3.3], "depreciation": [2.2],
        "working_capital_change": [0.0], "fcf": [11.25], "gross_profit": [55.0],
        "years": ["2024"],
        "statements": {
            "years_annual": ["2022", "2023", "2024", "2025"],
            "annual": {
                "profit_loss": {
                    "sales": [11.0, 12.1, 13.31, 14.641],
                    "operating profit": [1.65, 1.815, 1.9965, 2.19615],
                    "depreciation": [0.22, 0.242, 0.2662, 0.29282],
                    "interest": [0.0, 0.0, 0.0, 0.0],
                    "profit before tax": [1.65, 1.815, 1.9965, 2.19615],
                    "tax": [0.415, 0.456, 0.502, 0.552],
                    "net profit": [1.23, 1.35, 1.49, 1.64],
                },
                "balance_sheet": {
                    "equity capital": [8.0, 8.0, 8.0, 8.0],
                    "reserves": [0.0, 0.0, 0.0, 0.0],
                    "borrowings": [0.0, 0.0, 0.0, 0.0],
                    "fixed assets": [5.0, 5.0, 5.0, 5.0],
                    "trade payables": [0.0, 0.0, 0.0, 0.0],
                    "inventories": [0.0, 0.0, 0.0, 0.0],
                    "trade receivables": [0.0, 0.0, 0.0, 0.0],
                    "cash equivalents": [2.0, 2.0, 2.0, 2.0],
                },
                "cash_flow": {
                    "fixed assets purchased": [0.33, 0.363, 0.3993, 0.43923],
                },
            },
            "ratios": {},
        },
    }
    dcf_res = run_dcf_valuation(fin, {}, 0.04, None)
    excel_bytes = export_dcf_excel(dcf_res, fin)

    assert isinstance(excel_bytes, bytes)
    assert excel_bytes[:4] == b'PK\x03\x04', "Not a valid ZIP/XLSX file"

    wb2 = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert len(wb2.sheetnames) >= 1
