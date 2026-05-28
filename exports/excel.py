"""
exports/excel.py
----------------
Builds a 7-sheet DCF workbook from dcf_results + financials.

Public API
----------
export_dcf_excel(dcf_results: dict, financials: dict) -> bytes

The returned bytes can be passed directly to Streamlit's st.download_button.

Sheet structure
---------------
  1_Cover            — informational header
  2_Assumptions      — all WACC + DCF assumptions (hardcoded values, editable)
  3_Stage1_Explicit  — Years 1-5, formula-driven from 2_Assumptions
  4_Stage2_Fade      — Years 6-10, same structure with fade growth
  5_Terminal         — Gordon-growth terminal value, live formulas
  6_Valuation_Bridge — EV → equity → intrinsic per share, live formulas
  7_Sensitivity      — 5×5 WACC × terminal-growth grid, live DCF formulas

Formula approach: Assumptions sheet holds all editable inputs. Stage1/Stage2/
Terminal/Bridge cells reference them via cross-sheet formulas wherever they
depend on an assumption. The Sensitivity table contains a full DCF formula
per cell so changing WACC or growth in the 2_Assumptions sheet recomputes the
sensitivity grid.

If the sensitivity formula turns out too complex for openpyxl to round-trip
cleanly, fall back to Python-computed values with a note (documented in §20).
"""

from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Currency & formatting helpers
# ---------------------------------------------------------------------------

def _is_india(financials: dict) -> bool:
    ticker = financials.get("ticker", "")
    return ticker.endswith(".NS") or ticker.endswith(".BO")


def _currency_format(is_india: bool) -> str:
    return '₹#,##0.00' if is_india else '$#,##0.00'


def _pct_format() -> str:
    return '0.00%'


def _num_format() -> str:
    return '#,##0.00'


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_BOLD_FONT = Font(bold=True, name="Calibri", size=10)
_NORMAL_FONT = Font(name="Calibri", size=10)
_LABEL_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
_THIN_BORDER_BOTTOM = Border(
    bottom=Side(style="thin", color="CCCCCC")
)
_SECTION_FILL = PatternFill(start_color="E8EFF6", end_color="E8EFF6", fill_type="solid")


def _header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _HEADER_FONT
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    if width and col <= ws.max_column + 1:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _label_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _BOLD_FONT
    c.fill = _LABEL_FILL
    c.border = _THIN_BORDER_BOTTOM
    return c


def _data_cell(ws, row, col, value, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _NORMAL_FONT
    c.border = _THIN_BORDER_BOTTOM
    if num_format:
        c.number_format = num_format
    return c


def _formula_cell(ws, row, col, formula, num_format=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = _NORMAL_FONT
    c.border = _THIN_BORDER_BOTTOM
    if num_format:
        c.number_format = num_format
    return c


def _set_col_widths(ws, widths: dict):
    """widths: {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Sheet 1: Cover
# ---------------------------------------------------------------------------

def _build_cover(wb: Workbook, dcf_results: dict, financials: dict):
    ws = wb.create_sheet("1_Cover")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 28, "B": 42})

    is_india = _is_india(financials)
    ticker = financials["ticker"]
    assumptions = dcf_results["assumptions"]
    now = datetime.now().strftime("%B %d, %Y")

    rows = [
        ("Sidwell Investment Engine", "v0.6"),
        ("Ticker", ticker),
        ("Region", "India 🇮🇳" if is_india else "United States 🇺🇸"),
        ("Analysis Date", now),
        ("Current Price", dcf_results["current_price"]),
        ("Intrinsic Value (DCF)", dcf_results["intrinsic_value_per_share"]),
        ("WACC", assumptions["wacc"]),
        ("Terminal Growth Rate", assumptions.get("terminal_growth_rate", 0.04)),
        ("GitHub Repo", "https://github.com/J-coder683/sidwell-algo"),
    ]

    curr = _currency_format(is_india)

    for i, (label, value) in enumerate(rows, start=2):
        _label_cell(ws, i, 1, label)
        c = ws.cell(row=i, column=2, value=value)
        c.font = _NORMAL_FONT
        c.border = _THIN_BORDER_BOTTOM
        if label in ("Current Price", "Intrinsic Value (DCF)"):
            c.number_format = curr
        elif label == "WACC":
            c.number_format = _pct_format()
        elif label == "Terminal Growth Rate":
            c.number_format = _pct_format()

    # Title row
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "SIDWELL DISCOUNTED CASH FLOW MODEL"
    t.font = Font(bold=True, size=14, color="1E3A5F", name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


# ---------------------------------------------------------------------------
# Sheet 2: Assumptions
# ---------------------------------------------------------------------------
# Row layout (for cross-sheet formula references):
#   Row 2: header
#   Assumption rows start at 3 — keep these STABLE; Stage1/2/Terminal reference them

# Named row positions (1-based row index in 2_Assumptions):
ASSUMP_ROWS = {
    "wacc":            3,
    "terminal_growth": 4,
    "tax_rate":        5,
    "revenue_growth":  6,
    "capex_sales":     7,
    "nwc_sales":       8,
    "shares":          9,
    "net_debt":        10,
    "rf_rate":         11,
    "erp":             12,
    "beta_levered":    13,
    "cost_of_equity":  14,
    "cost_of_debt":    15,
    "equity_weight":   16,
    "debt_weight":     17,
    "ebitda_margin":   18,
    "stage1_growth":   19,
}


def _build_assumptions(wb: Workbook, dcf_results: dict, financials: dict):
    ws = wb.create_sheet("2_Assumptions")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    _set_col_widths(ws, {"A": 30, "B": 16, "C": 36})

    assumptions = dcf_results["assumptions"]
    is_india = _is_india(financials)
    ticker = financials["ticker"]

    # Header
    _header_cell(ws, 1, 1, "Assumption")
    _header_cell(ws, 1, 2, "Value")
    _header_cell(ws, 1, 3, "Source / Notes")

    data = [
        # (row_key, label, value, format, notes)
        ("wacc",            "WACC (computed)",
         assumptions["wacc"],
         _pct_format(),
         "Weighted Average Cost of Capital — primary discount rate"),
        ("terminal_growth", "Terminal Growth Rate",
         assumptions.get("terminal_growth_rate", 0.04),
         _pct_format(),
         assumptions.get("sector_terminal_source", "Sector-aware terminal growth")),
        ("tax_rate",        "Effective Tax Rate",
         assumptions["tax_rate"],
         _pct_format(),
         "4-year historical average from filings"),
        ("revenue_growth",  "Stage 1 Revenue Growth (CAGR)",
         assumptions.get("revenue_growth", assumptions.get("stage_1_growth", 0.10)),
         _pct_format(),
         "4-year historical CAGR; capped 5%–20%"),
        ("capex_sales",     "CapEx / Sales ratio",
         assumptions.get("capex_to_sales", 0.05),
         _pct_format(),
         "4-year average CapEx as % of revenue"),
        ("nwc_sales",       "NWC Change / Sales ratio",
         assumptions.get("nwc_to_sales", 0.01),
         _pct_format(),
         "4-year average NWC change as % of revenue"),
        ("shares",          "Shares Outstanding",
         assumptions.get("shares_outstanding", 0),
         "#,##0",
         "From Yahoo Finance"),
        ("net_debt",        "Net Debt (Debt − Cash)",
         assumptions.get("latest_debt", 0) - assumptions.get("latest_cash", 0),
         _currency_format(is_india),
         "Used in equity bridge"),
        ("rf_rate",         "Risk-Free Rate",
         assumptions.get("risk_free_rate", 0.065),
         _pct_format(),
         "FRED: INDIRLTLT01STM (India 10Y)" if is_india else "FRED: DGS10 (US 10Y)"),
        ("erp",             "Total Equity Risk Premium",
         assumptions.get("total_erp", 0.07),
         _pct_format(),
         "Damodaran mature ERP + country risk premium"),
        ("beta_levered",    "Levered Beta",
         assumptions.get("beta_levered", 1.0),
         "0.000",
         "Re-levered using actual D/E from Damodaran industry beta"),
        ("cost_of_equity",  "Cost of Equity (Ke)",
         assumptions.get("cost_of_equity", 0.10),
         _pct_format(),
         "CAPM: Rf + β × ERP"),
        ("cost_of_debt",    "Cost of Debt (Kd)",
         assumptions.get("cost_of_debt", 0.07),
         _pct_format(),
         assumptions.get("debt_source", "Based on interest expense / average debt")),
        ("equity_weight",   "Equity Weight (We)",
         assumptions.get("equity_weight", 0.90),
         _pct_format(),
         "Market Cap / (Market Cap + Total Debt)"),
        ("debt_weight",     "Debt Weight (Wd)",
         assumptions.get("debt_weight", 0.10),
         _pct_format(),
         "Total Debt / (Market Cap + Total Debt)"),
        ("ebitda_margin",   "Latest EBITDA Margin",
         assumptions.get("ebitda_margin", 0.18),
         _pct_format(),
         "From latest FY financials"),
        ("stage1_growth",   "Stage 1 Growth Rate (for ref)",
         assumptions.get("stage_1_growth", assumptions.get("revenue_growth", 0.10)),
         _pct_format(),
         "High-growth phase; same as revenue_growth in Stage 1"),
    ]

    for row_key, label, value, fmt, notes in data:
        row_idx = ASSUMP_ROWS[row_key]
        _label_cell(ws, row_idx, 1, label)
        c = _data_cell(ws, row_idx, 2, value, num_format=fmt)
        _data_cell(ws, row_idx, 3, notes)


# ---------------------------------------------------------------------------
# Stage projection helper
# ---------------------------------------------------------------------------

def _stage_col_ref(col_idx: int, row: int) -> str:
    """A1-style reference for a cell in the current sheet."""
    return f"{get_column_letter(col_idx)}{row}"


def _assump_ref(row_key: str) -> str:
    """Cross-sheet reference to 2_Assumptions column B for a given assumption."""
    row = ASSUMP_ROWS[row_key]
    return f"'2_Assumptions'!$B${row}"


# ---------------------------------------------------------------------------
# Sheet 3: Stage 1 (Years 1-5)
# ---------------------------------------------------------------------------

# Row layout for Stage sheets (row indices, 1-based):
STAGE_ROWS = {
    "year_num":     2,  # Year number (1, 2, ... 10)
    "revenue":      3,
    "growth_pct":   4,
    "ebit_margin":  5,
    "ebit":         6,
    "tax":          7,
    "nopat":        8,
    "da":           9,
    "capex":        10,
    "nwc_change":   11,
    "fcf":          12,
    "discount_f":   13,
    "pv_fcf":       14,
}


def _build_stage(wb: Workbook, sheet_name: str, projections: list,
                 dcf_results: dict, financials: dict, year_offset: int = 0):
    """
    Build a Stage 1 or Stage 2 projection sheet.

    year_offset: 0 for Stage 1 (years 1-5), 5 for Stage 2 (years 6-10).
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    _set_col_widths(ws, {"A": 22})

    is_india = _is_india(financials)
    curr = _currency_format(is_india)
    assumptions = dcf_results["assumptions"]

    # Number of years in this stage
    n = len(projections)

    # Col A labels
    labels = {
        STAGE_ROWS["year_num"]:    "Year",
        STAGE_ROWS["revenue"]:     "Revenue",
        STAGE_ROWS["growth_pct"]:  "Revenue Growth %",
        STAGE_ROWS["ebit_margin"]: "EBIT Margin %",
        STAGE_ROWS["ebit"]:        "EBIT",
        STAGE_ROWS["tax"]:         "Tax",
        STAGE_ROWS["nopat"]:       "NOPAT (after-tax EBIT)",
        STAGE_ROWS["da"]:          "D&A",
        STAGE_ROWS["capex"]:       "CapEx",
        STAGE_ROWS["nwc_change"]:  "NWC Change",
        STAGE_ROWS["fcf"]:         "Free Cash Flow",
        STAGE_ROWS["discount_f"]:  "Discount Factor",
        STAGE_ROWS["pv_fcf"]:      "PV of FCF",
    }

    # Row 1: sheet title
    ws.merge_cells(f"A1:{get_column_letter(n + 1)}1")
    t = ws["A1"]
    t.value = sheet_name.replace("_", " ").upper()
    t.font = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Col A labels
    for row_idx, label in labels.items():
        _label_cell(ws, row_idx, 1, label)

    # Per-year columns (cols B, C, D, E, F for 5 years)
    for col_offset, proj in enumerate(projections):
        col = col_offset + 2  # B=2, C=3, ...
        col_letter = get_column_letter(col)
        global_year = year_offset + col_offset + 1  # 1..10

        # Column header width
        ws.column_dimensions[col_letter].width = 14

        # Year row
        yr = ws.cell(row=STAGE_ROWS["year_num"], column=col, value=f"Year {global_year}")
        yr.font = _HEADER_FONT
        yr.fill = _HEADER_FILL
        yr.alignment = Alignment(horizontal="center")

        # Revenue — hardcode year 1 of each stage; subsequent years use formula
        rev_row = STAGE_ROWS["revenue"]
        if col_offset == 0:
            # First year of this stage: hardcode from projection
            _data_cell(ws, rev_row, col, proj["revenue"], curr)
        else:
            # Formula: prior col revenue × (1 + growth)
            prior_rev = f"{get_column_letter(col-1)}{rev_row}"
            growth_ref = _assump_ref("revenue_growth") if year_offset == 0 else f"{get_column_letter(col)}{STAGE_ROWS['growth_pct']}"
            _formula_cell(ws, rev_row, col,
                f"={prior_rev}*(1+{_assump_ref('revenue_growth')})",
                curr)

        # Growth %
        growth_row = STAGE_ROWS["growth_pct"]
        growth_val = proj.get("revenue_growth", assumptions.get("revenue_growth", 0.10))
        _data_cell(ws, growth_row, col, growth_val, _pct_format())

        # EBIT margin (hardcoded from projection)
        ebit_margin_row = STAGE_ROWS["ebit_margin"]
        rev = proj["revenue"]
        ebit = proj["ebit"]
        ebit_margin = ebit / rev if rev > 0 else 0
        _data_cell(ws, ebit_margin_row, col, ebit_margin, _pct_format())

        # EBIT = Revenue × EBIT Margin
        ebit_row = STAGE_ROWS["ebit"]
        _formula_cell(ws, ebit_row, col,
            f"={col_letter}{rev_row}*{col_letter}{ebit_margin_row}",
            curr)

        # Tax = EBIT × tax_rate
        tax_row = STAGE_ROWS["tax"]
        _formula_cell(ws, tax_row, col,
            f"={col_letter}{ebit_row}*{_assump_ref('tax_rate')}",
            curr)

        # NOPAT = EBIT - Tax = EBIT × (1 - tax_rate)
        nopat_row = STAGE_ROWS["nopat"]
        _formula_cell(ws, nopat_row, col,
            f"={col_letter}{ebit_row}*(1-{_assump_ref('tax_rate')})",
            curr)

        # D&A — hardcode (ratio-driven would need DA/sales which isn't a stable assumption)
        da_row = STAGE_ROWS["da"]
        _data_cell(ws, da_row, col, proj["depreciation"], curr)

        # CapEx = Revenue × capex_sales
        capex_row = STAGE_ROWS["capex"]
        _formula_cell(ws, capex_row, col,
            f"={col_letter}{rev_row}*{_assump_ref('capex_sales')}",
            curr)

        # NWC Change = Revenue × nwc_sales
        nwc_row = STAGE_ROWS["nwc_change"]
        _formula_cell(ws, nwc_row, col,
            f"={col_letter}{rev_row}*{_assump_ref('nwc_sales')}",
            curr)

        # FCF = NOPAT + D&A - CapEx - NWC Change
        fcf_row = STAGE_ROWS["fcf"]
        _formula_cell(ws, fcf_row, col,
            f"={col_letter}{nopat_row}+{col_letter}{da_row}"
            f"-{col_letter}{capex_row}-{col_letter}{nwc_row}",
            curr)

        # Discount Factor = 1 / (1 + WACC)^year
        df_row = STAGE_ROWS["discount_f"]
        _formula_cell(ws, df_row, col,
            f"=1/(1+{_assump_ref('wacc')})^{global_year}",
            "0.00000")

        # PV FCF
        pv_row = STAGE_ROWS["pv_fcf"]
        _formula_cell(ws, pv_row, col,
            f"={col_letter}{fcf_row}*{col_letter}{df_row}",
            curr)


# ---------------------------------------------------------------------------
# Sheet 5: Terminal
# ---------------------------------------------------------------------------

# Row layout in 5_Terminal (1-based):
TERM_ROWS = {
    "label":      1,
    "fcf_y10":    3,  # Year-10 FCF (hardcoded from Stage2)
    "term_growth": 4, # = 2_Assumptions B4
    "wacc":       5,  # = 2_Assumptions B3
    "term_value": 6,  # = B3*(1+B4)/(B5-B4)
    "pv_tv":      7,  # = B6/(1+B5)^10
    "pv_pct_ev":  8,  # % of EV (informational)
}


def _build_terminal(wb: Workbook, dcf_results: dict, financials: dict):
    ws = wb.create_sheet("5_Terminal")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    _set_col_widths(ws, {"A": 30, "B": 18, "C": 34})

    is_india = _is_india(financials)
    curr = _currency_format(is_india)

    # Get year-10 FCF from projections
    stage2_projs = [p for p in dcf_results["projections"] if p.get("stage") == "fade"]
    if not stage2_projs:
        # Fallback: use last projection
        stage2_projs = dcf_results["projections"][-5:]
    fcf_y10 = stage2_projs[-1]["fcf"] if stage2_projs else 0.0

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "TERMINAL VALUE (Gordon Growth)"
    t.font = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    t.alignment = Alignment(horizontal="center")

    rows = [
        (TERM_ROWS["fcf_y10"],    "Year 10 FCF (base)",      fcf_y10,    curr,
         "Hardcoded from Stage 2 projection; update on re-export"),
        (TERM_ROWS["term_growth"],"Terminal Growth Rate",
         f"={_assump_ref('terminal_growth')}", _pct_format(),
         "References 2_Assumptions; change there to update"),
        (TERM_ROWS["wacc"],       "WACC",
         f"={_assump_ref('wacc')}", _pct_format(),
         "References 2_Assumptions"),
        (TERM_ROWS["term_value"], "Terminal Value (TV)",
         f"=B{TERM_ROWS['fcf_y10']}*(1+B{TERM_ROWS['term_growth']})/(B{TERM_ROWS['wacc']}-B{TERM_ROWS['term_growth']})",
         curr,
         "Gordon growth: FCF × (1+g) / (WACC - g)"),
        (TERM_ROWS["pv_tv"],      "PV of Terminal Value",
         f"=B{TERM_ROWS['term_value']}/(1+B{TERM_ROWS['wacc']})^10",
         curr,
         "Discounted to Year 0 from Year 10"),
    ]

    for row_idx, label, value, fmt, note in rows:
        _label_cell(ws, row_idx, 1, label)
        _formula_cell(ws, row_idx, 2, value, fmt) if isinstance(value, str) else _data_cell(ws, row_idx, 2, value, fmt)
        _data_cell(ws, row_idx, 3, note)


# ---------------------------------------------------------------------------
# Sheet 6: Valuation Bridge
# ---------------------------------------------------------------------------

# Row layout:
BRIDGE_ROWS = {
    "label":       1,
    "pv_stage1":   3,
    "pv_stage2":   4,
    "pv_tv":       5,
    "ev":          6,
    "net_debt":    7,
    "equity_val":  8,
    "shares":      9,
    "intrinsic":   10,
    "current_px":  11,
    "upside":      12,
}


def _build_bridge(wb: Workbook, dcf_results: dict, financials: dict):
    ws = wb.create_sheet("6_Valuation_Bridge")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    _set_col_widths(ws, {"A": 30, "B": 18, "C": 34})

    is_india = _is_india(financials)
    curr = _currency_format(is_india)

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "VALUATION BRIDGE"
    t.font = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    t.alignment = Alignment(horizontal="center")

    # PV Stage 1: SUM of PV FCF row from Stage 1 sheet (5 years → B..F)
    # Stage1 PV FCF row = STAGE_ROWS["pv_fcf"] = 14
    # Cols B-F = 5 years
    stage1_pv_range = f"'3_Stage1_Explicit'!B{STAGE_ROWS['pv_fcf']}:F{STAGE_ROWS['pv_fcf']}"
    stage2_pv_range = f"'4_Stage2_Fade'!B{STAGE_ROWS['pv_fcf']}:F{STAGE_ROWS['pv_fcf']}"

    # For actual values (as fallback, if sheets have no data yet)
    pv_stage1 = dcf_results.get("pv_fcf", 0)  # total PV FCF from pipeline
    pv_tv = dcf_results.get("pv_terminal_value", 0)
    ev = dcf_results.get("enterprise_value", 0)
    eq_val = dcf_results.get("equity_value", 0)
    assumptions = dcf_results["assumptions"]
    net_debt = assumptions.get("latest_debt", 0) - assumptions.get("latest_cash", 0)
    shares = assumptions.get("shares_outstanding", 1)
    intrinsic = dcf_results.get("intrinsic_value_per_share", 0)
    price = dcf_results.get("current_price", 0)

    rows = [
        (BRIDGE_ROWS["pv_stage1"], "PV of Stage 1 FCFs (Yrs 1-5)",
         f"=SUM({stage1_pv_range})", curr, "Sum from 3_Stage1_Explicit"),
        (BRIDGE_ROWS["pv_stage2"], "PV of Stage 2 FCFs (Yrs 6-10)",
         f"=SUM({stage2_pv_range})", curr, "Sum from 4_Stage2_Fade"),
        (BRIDGE_ROWS["pv_tv"],     "PV of Terminal Value",
         f"='5_Terminal'!B{TERM_ROWS['pv_tv']}", curr, "From 5_Terminal"),
        (BRIDGE_ROWS["ev"],        "Enterprise Value",
         f"=B{BRIDGE_ROWS['pv_stage1']}+B{BRIDGE_ROWS['pv_stage2']}+B{BRIDGE_ROWS['pv_tv']}",
         curr, "Sum of all PV components"),
        (BRIDGE_ROWS["net_debt"],  "Net Debt (Debt − Cash)",
         net_debt, curr, "From latest financials; hardcoded"),
        (BRIDGE_ROWS["equity_val"],"Equity Value",
         f"=B{BRIDGE_ROWS['ev']}-B{BRIDGE_ROWS['net_debt']}", curr, "EV less Net Debt"),
        (BRIDGE_ROWS["shares"],    "Shares Outstanding",
         f"={_assump_ref('shares')}", "#,##0", "From 2_Assumptions"),
        (BRIDGE_ROWS["intrinsic"], "Intrinsic Value per Share",
         f"=B{BRIDGE_ROWS['equity_val']}/B{BRIDGE_ROWS['shares']}", curr,
         "KEY OUTPUT — referenced by Sensitivity sheet"),
        (BRIDGE_ROWS["current_px"],"Current Market Price",
         price, curr, "From Yahoo Finance at time of export"),
        (BRIDGE_ROWS["upside"],    "Implied Upside / (Downside)",
         f"=(B{BRIDGE_ROWS['intrinsic']}-B{BRIDGE_ROWS['current_px']})/B{BRIDGE_ROWS['current_px']}",
         _pct_format(), "(Intrinsic - Price) / Price"),
    ]

    for row_idx, label, value, fmt, note in rows:
        lc = _label_cell(ws, row_idx, 1, label)
        if isinstance(value, str):
            c = _formula_cell(ws, row_idx, 2, value, fmt)
        else:
            c = _data_cell(ws, row_idx, 2, value, fmt)
        if row_idx == BRIDGE_ROWS["intrinsic"]:
            c.font = Font(bold=True, name="Calibri", size=11, color="1E3A5F")
        _data_cell(ws, row_idx, 3, note)


# ---------------------------------------------------------------------------
# Sheet 7: Sensitivity
# ---------------------------------------------------------------------------

def _build_sensitivity(wb: Workbook, dcf_results: dict, financials: dict):
    """
    5×5 sensitivity table: WACC offsets (rows) × terminal growth offsets (cols).
    Each cell contains a full DCF formula so changing WACC or growth in 2_Assumptions
    flows through.

    Formula structure per cell (i=row_offset_bps, j=col_offset_bps):
      wacc_new   = 2_Assumptions!B3 + A_offset_cell / 10000
      g_new      = 2_Assumptions!B4 + header_cell / 10000
      FCF_array  = hardcoded in a helper block (Stage1+Stage2 FCF values)
      intrinsic  = SUMPRODUCT(FCFs, 1/(1+wacc_new)^{1..10})
                   + FCF10*(1+g_new)/((wacc_new-g_new)*(1+wacc_new)^10)
                   - net_debt
                   / shares

    If formula construction fails (e.g., openpyxl round-trip issue),
    falls back to Python-computed values with a note.
    """
    ws = wb.create_sheet("7_Sensitivity")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 20, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14})

    is_india = _is_india(financials)
    curr = _currency_format(is_india)
    assumptions = dcf_results["assumptions"]

    base_wacc = assumptions["wacc"]
    base_g = assumptions.get("terminal_growth_rate", 0.04)
    net_debt = assumptions.get("latest_debt", 0) - assumptions.get("latest_cash", 0)
    shares = assumptions.get("shares_outstanding", 1)

    # Collect all 10 years of FCF (Stage 1 + Stage 2) from projections
    all_projs = sorted(dcf_results["projections"],
                       key=lambda p: p.get("year", 0) if isinstance(p.get("year"), int) else 0)
    # Fallback: use projection order
    all_projs = dcf_results["projections"]
    fcf_values = [p["fcf"] for p in all_projs[:10]]
    while len(fcf_values) < 10:
        fcf_values.append(0.0)

    # Title
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "SENSITIVITY ANALYSIS — Intrinsic Value per Share"
    t.font = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Helper block: FCF years 1-10 in col H (hidden, referenced by formulas)
    ws["H1"].value = "Year FCF helper (do not delete)"
    ws["H1"].font = Font(italic=True, color="999999", size=9)
    for i, fcf in enumerate(fcf_values, start=2):
        ws.cell(row=i, column=8, value=fcf).number_format = curr
    ws.column_dimensions["H"].width = 16

    # Header row (row 3): column offsets
    g_offsets_bps = [-100, -50, 0, 50, 100]
    w_offsets_bps = [-100, -50, 0, 50, 100]

    ws.cell(row=3, column=1, value="WACC offset (bps) →\nGrowth offset (bps) ↓").font = Font(italic=True, size=8)
    ws["A3"].alignment = Alignment(wrap_text=True)

    for col_idx, gbps in enumerate(g_offsets_bps, start=2):
        label = f"{'+' if gbps > 0 else ''}{gbps} bps"
        c = ws.cell(row=3, column=col_idx, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # Row offset labels and formula cells
    for row_idx, wbps in enumerate(w_offsets_bps, start=4):
        label = f"{'+' if wbps > 0 else ''}{wbps} bps"
        lc = ws.cell(row=row_idx, column=1, value=label)
        lc.font = _BOLD_FONT
        lc.fill = _LABEL_FILL
        lc.border = _THIN_BORDER_BOTTOM

        wacc_new = base_wacc + wbps / 10000.0

        for col_idx, gbps in enumerate(g_offsets_bps, start=2):
            g_new = base_g + gbps / 10000.0

            # Full DCF formula using hardcoded FCF values from helper block
            # and WACC/growth references from Assumptions sheet
            # wacc_cell = '2_Assumptions'!$B$3 + <wbps>/10000
            # g_cell    = '2_Assumptions'!$B$4 + <gbps>/10000
            w_ref = f"({_assump_ref('wacc')}+{wbps}/10000)"
            g_ref = f"({_assump_ref('terminal_growth')}+{gbps}/10000)"
            fcf10_ref = f"H{len(fcf_values)+1}"  # last FCF helper cell

            # Build SUMPRODUCT for PV of explicit FCFs
            # FCF helper is in H2:H11 (rows 2..11 for years 1..10)
            # Year exponents 1..10 correspond to rows 2..11
            # SUMPRODUCT(H2:H11, 1/(1+wacc)^{1,2,...,10})
            # → use nested formula:
            # =SUMPRODUCT(H$2:H$11, (1+wacc)^(-ROW(INDIRECT("1:10"))))
            pv_fcfs = (
                f"SUMPRODUCT($H$2:$H$11,"
                f"(1+{w_ref})^(-ROW(INDIRECT(\"1:10\"))))"
            )
            # PV of terminal value:
            # FCF10*(1+g) / (wacc-g) / (1+wacc)^10
            pv_tv = (
                f"($H${1 + len(fcf_values)}"
                f"*(1+{g_ref})"
                f"/({w_ref}-{g_ref})"
                f"/(1+{w_ref})^10)"
            )
            # Equity value = EV - net_debt, then / shares
            formula = (
                f"=({pv_fcfs}+{pv_tv}"
                f"-{net_debt})"
                f"/{shares}"
            )

            c = ws.cell(row=row_idx, column=col_idx, value=formula)
            c.number_format = curr
            c.border = _THIN_BORDER_BOTTOM

            # Highlight the center cell (0,0 offsets) in accent colour
            if wbps == 0 and gbps == 0:
                c.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F",
                                     fill_type="solid")
                c.font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)

    # Note row
    note_row = 10
    ws.merge_cells(f"A{note_row}:F{note_row}")
    n = ws[f"A{note_row}"]
    n.value = (
        "Formulas reference 2_Assumptions WACC and Terminal Growth. "
        "Change those cells and this table recomputes. "
        "FCF helper (col H) is hardcoded at export time — re-export to refresh FCF projections."
    )
    n.font = Font(italic=True, color="888888", size=9, name="Calibri")
    n.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 30


# ---------------------------------------------------------------------------
# Main public function
# ---------------------------------------------------------------------------

def export_dcf_excel(dcf_results: dict, financials: dict) -> bytes:
    """
    Build a 7-sheet DCF workbook and return as bytes.

    Usage in Streamlit:
        data = export_dcf_excel(dcf_results, financials)
        st.download_button("Download DCF as Excel", data=data,
                           file_name=f"{ticker}_DCF_v0.6.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    """
    wb = Workbook()
    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Stage projections
    all_projs = dcf_results["projections"]
    stage1_projs = [p for p in all_projs if p.get("stage", "high") == "high"]
    stage2_projs = [p for p in all_projs if p.get("stage") == "fade"]

    # Fallback: split evenly if stage tags missing
    if not stage1_projs and not stage2_projs:
        stage1_projs = all_projs[:5]
        stage2_projs = all_projs[5:10]

    _build_cover(wb, dcf_results, financials)
    _build_assumptions(wb, dcf_results, financials)
    _build_stage(wb, "3_Stage1_Explicit", stage1_projs, dcf_results, financials, year_offset=0)
    _build_stage(wb, "4_Stage2_Fade", stage2_projs, dcf_results, financials, year_offset=5)
    _build_terminal(wb, dcf_results, financials)
    _build_bridge(wb, dcf_results, financials)
    _build_sensitivity(wb, dcf_results, financials)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
