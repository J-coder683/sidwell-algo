import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from sidwell.render.formats import Formats
from sidwell.ajp.schema import AJP
from sidwell.ajp.loader import AJPLoader

class WorkbookRenderer:
    def __init__(self, engine_results: Dict[str, Any], ajp: AJP):
        self.results = engine_results
        self.ajp = ajp
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active) # Remove default sheet
        self._ticker = str(self.results.get("meta", {}).get("ticker", self.ajp.meta.ticker)).upper()
        _is_india = self._ticker.endswith(".NS") or self._ticker.endswith(".BO")
        self._cur  = "Rs" if _is_india else "USD"
        self._unit = f"{self._cur} mm"
        
    def _create_sheet(self, title: str):
        ws = self.wb.create_sheet(title)
        # Apply standard column widths
        ws.column_dimensions['A'].width = 40
        for i in range(2, 20):
            ws.column_dimensions[get_column_letter(i)].width = 15
        return ws
        
    def _write_header(self, ws, row: int, col: int, text: str):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Formats.FONT_HEADER
        cell.fill = Formats.FILL_HEADER

    def _proj_header(self, ws, title):
        """Title + projection year columns (FY..E), banker header style."""
        c = ws.cell(row=2, column=2, value=title)
        c.font = Formats.FONT_HEADER
        c.fill = Formats.FILL_HEADER
        for i, y in enumerate(self.results["proj"]["years"]):
            hc = ws.cell(row=2, column=3 + i, value=self._fy_label(y, "E"))
            hc.font = Formats.FONT_HEADER
            hc.fill = Formats.FILL_HEADER
            hc.alignment = Formats.ALIGN_CENTER

    def _row(self, ws, r, label, values, fmt=None, bold=False):
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = Formats.FONT_BOLD if bold else Formats.FONT_NORMAL
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=3 + i, value=v)
            cell.number_format = fmt or Formats.FMT_NUMBER
        return r + 1

    def _kv(self, ws, r, label, value, fmt=None, blue=False):
        ws.cell(row=r, column=2, value=label).font = Formats.FONT_BOLD
        cell = ws.cell(row=r, column=3, value=value)
        cell.number_format = fmt or Formats.FMT_NUMBER
        if blue:
            cell.font = Formats.FONT_INPUT
        return r + 1

    def _av_header(self, ws, title):
        """Historical actual (FY..A) + projected (FY..E) year columns. Returns n_hist."""
        c = ws.cell(row=2, column=2, value=title)
        c.font = Formats.FONT_HEADER
        c.fill = Formats.FILL_HEADER
        h_years = self.results.get("hist", {}).get("years_annual", [])
        col = 3
        for y in h_years:
            hc = ws.cell(row=2, column=col, value=self._fy_label(y, "A"))
            hc.font = Formats.FONT_HEADER; hc.fill = Formats.FILL_HEADER; hc.alignment = Formats.ALIGN_CENTER
            col += 1
        for y in self.results["proj"]["years"]:
            hc = ws.cell(row=2, column=col, value=self._fy_label(y, "E"))
            hc.font = Formats.FONT_HEADER; hc.fill = Formats.FILL_HEADER; hc.alignment = Formats.ALIGN_CENTER
            col += 1
        return len(h_years)

    def _av_row(self, ws, r, label, hvals, pvals, n_hist, fmt=None, bold=False):
        """Row across historical (blue) + projected (black) columns. Historical
        cells that are missing (0/None from un-scraped years) are left blank."""
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = Formats.FONT_BOLD if bold else Formats.FONT_NORMAL
        c = 3
        for v in (hvals[-n_hist:] if n_hist else hvals):
            if v:  # skip 0/None → not-reported, leave blank
                cell = ws.cell(row=r, column=c, value=v)
                cell.number_format = fmt or Formats.FMT_NUMBER
                cell.font = Formats.FONT_INPUT  # blue = actual
            c += 1
        for v in pvals:
            cell = ws.cell(row=r, column=c, value=v)
            cell.number_format = fmt or Formats.FMT_NUMBER
            cell.font = Formats.FONT_NORMAL  # black = projected
            c += 1
        return r + 1
        
    def render(self) -> openpyxl.Workbook:
        """Renders the full workbook with 13 sheets and live formulas."""
        self.render_cover()
        self.render_drivers()
        self.render_assumptions()
        self.render_is()
        self.render_bs()
        self.render_cf()
        self.render_debt()
        self.render_fcf()
        self.render_wacc()
        self.render_terminal()
        self.render_bridge()
        self.render_sensitivity()
        self.render_sources()
        return self.wb

    def render_cover(self):
        ws = self._create_sheet("1_Cover")
        ws.cell(row=2, column=2, value="Sidwell DCF Valuation").font = Formats.FONT_BOLD
        ws.cell(row=4, column=2, value="Ticker:")
        ws.cell(row=4, column=3, value=self._ticker).font = Formats.FONT_BOLD
        ws.cell(row=5, column=2, value=f"Intrinsic Value ({self._cur}):")
        ws.cell(row=5, column=3, value="='11_Valuation_Bridge'!C13").font = Formats.FONT_LINK
        
    def render_drivers(self):
        ws = self._create_sheet("2_Drivers_Scenarios")
        self._write_header(ws, 2, 2, "Scenario Switch")
        ws.cell(row=3, column=2, value="Active Scenario:")
        
        active_cell = ws.cell(row=3, column=3, value=self.ajp.meta.scenario_active)
        active_cell.font = Formats.FONT_INPUT
        
    def render_assumptions(self):
        ws = self._create_sheet("3_Assumptions_Just")
        self._write_header(ws, 2, 2, "Input")
        self._write_header(ws, 2, 3, "Value")
        self._write_header(ws, 2, 4, "Notes / Justification")
        ws.column_dimensions['D'].width = 80
        
        row = 3
        for a in self.ajp.assumptions:
            ws.cell(row=row, column=2, value=a.driver_id)
            val_cell = ws.cell(row=row, column=3, value=a.value)
            val_cell.font = Formats.FONT_INPUT
            
            note_text = f"[{a.source_type}] {a.rationale}"
            if a.verify_flag or a.confidence in ["LOW", "UNVERIFIED"]:
                note_text += f" [VERIFY: {a.verify_flag or 'UNVERIFIED'}]"
                val_cell.fill = Formats.FILL_FLAGGED
                
            ws.cell(row=row, column=4, value=note_text)
            row += 1

        # When no Gemini AJP is supplied, show the engine's data-derived drivers
        # and their historical basis so the assumptions are still justified.
        if not self.ajp.assumptions:
            au = self.results["proj"].get("assumptions_used", {})
            basis = {
                "stage1_revenue_growth": "Historical 10-yr revenue CAGR",
                "hist_revenue_cagr": "Historical 10-yr revenue CAGR",
                "terminal_growth": "Long-run nominal GDP proxy (fallback 2%)",
                "ebit_margin_start": "Latest historical EBIT margin",
                "ebit_margin_target": "Held at latest margin (no AJP expansion view)",
                "tax_rate": "Avg effective tax (tax / PBT) over history",
                "capex_pct_sales": "Derived CapEx (from Net Block & D&A) as % of sales",
                "da_rate_on_block": "Effective depreciation rate on net fixed-asset base",
                "dso_days": "Historical debtor days",
                "dio_days": "Historical inventory days",
                "dpo_days": "Historical payable days",
            }
            pct_keys = {"stage1_revenue_growth", "hist_revenue_cagr", "terminal_growth",
                        "ebit_margin_start", "ebit_margin_target", "tax_rate",
                        "capex_pct_sales", "da_rate_on_block"}
            for k, v in au.items():
                ws.cell(row=row, column=2, value=k)
                vc = ws.cell(row=row, column=3, value=v)
                vc.font = Formats.FONT_INPUT
                vc.number_format = Formats.FMT_PERCENT if k in pct_keys else '0.0'
                ws.cell(row=row, column=4, value=f"[ENGINE-EST] {basis.get(k, 'Derived from historical data')}")
                row += 1

    @staticmethod
    def _fy_label(raw, suffix):
        """'Mar 2025' -> 'FY2025A'. Extracts a real 4-digit year (handles noisy
        labels like 'Mar 2024\\n...15m' which must NOT become FY2415)."""
        import re
        s = str(raw)
        if s.startswith("FY"):
            return s
        m = re.search(r"(19|20)\d{2}", s)
        return f"FY{m.group(0)}{suffix}" if m else (s.strip()[:7] + suffix)

    def render_is(self):
        """Formula-driven Income Statement: projection DRIVER rows are blue editable
        inputs; Revenue/EBIT/NOPAT/CapEx/D&A are live formulas off those drivers and
        the prior column. All references local to this sheet (robust)."""
        from openpyxl.utils import get_column_letter as L
        F = Formats
        ws = self._create_sheet("4_Income_Statement")
        hist = self.results.get("hist", {})
        proj = self.results["proj"]
        au = proj.get("assumptions_used", {})
        h_is = hist.get("is", {})
        h_cf = hist.get("cf", {})
        h_years = hist.get("years_annual", [])
        n = len(h_years)
        hist_rev = h_is.get("sales") or h_is.get("revenue") or []
        hist_cogs = h_is.get("cogs") or []
        hist_ebit = h_is.get("operating_profit") or []
        hist_np = h_is.get("net_profit") or []
        hist_da = h_is.get("depreciation") or []
        hist_capex = h_cf.get("derived_capex") or [abs(c) for c in (h_cf.get("fixed_assets_purchased") or [])]
        hist_ppe = (hist.get("bs", {}) or {}).get("fixed_assets") or []
        hpe = (hist_ppe or [])[-n:] if n else []   # last-n net block values for D&A formula
        years = proj["years"]
        ny = len(years)

        t = ws.cell(row=2, column=2, value=f"Income Statement ({self._unit})")
        t.font = F.FONT_HEADER; t.fill = F.FILL_HEADER
        col = 3
        for y in h_years:
            c = ws.cell(row=2, column=col, value=self._fy_label(y, "A"))
            c.font = F.FONT_HEADER; c.fill = F.FILL_HEADER; c.alignment = F.ALIGN_CENTER
            col += 1
        pc0 = col  # first projection column
        for y in years:
            c = ws.cell(row=2, column=col, value=self._fy_label(y, "E"))
            c.font = F.FONT_HEADER; c.fill = F.FILL_HEADER; c.alignment = F.ALIGN_CENTER
            col += 1

        R_G, R_REV, R_COGSP, R_COGS, R_MG, R_EBIT, R_TAX, R_NOP, R_CXP, R_CX, R_DAP, R_DA, R_INT, R_NP = 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16
        for rr, lab in [(R_G, "Revenue growth %"), (R_REV, "Revenue"), (R_COGSP, "COGS % sales"), (R_COGS, "COGS"), 
                        (R_MG, "EBIT margin %"), (R_EBIT, "EBIT"), (R_TAX, "Tax rate %"), (R_NOP, "NOPAT"),
                        (R_CXP, "CapEx % sales"), (R_CX, "CapEx"), (R_DAP, "D&A rate on PP&E %"),
                        (R_DA, "D&A"), (R_INT, "Interest Expense"), (R_NP, "Net Profit")]:
            ws.cell(row=rr, column=2, value=lab).font = F.FONT_BOLD

        def put_hist(row, vals, fmt=F.FMT_NUMBER):
            for i, v in enumerate(vals[-n:] if n else []):
                if not v:   # missing/un-scraped → leave blank
                    continue
                c = ws.cell(row=row, column=3 + i, value=v)
                c.number_format = fmt; c.font = F.FONT_INPUT
        put_hist(R_REV, hist_rev); put_hist(R_COGS, hist_cogs); put_hist(R_EBIT, hist_ebit)
        put_hist(R_NP, hist_np); put_hist(R_DA, hist_da); put_hist(R_CX, hist_capex)
        hist_int = h_is.get("interest") or []
        put_hist(R_INT, hist_int)

        # Historical driver rows (ACTUALS) — LIVE Excel formulas referencing the
        # line-item rows already in this sheet (same pattern as projection columns).
        # Tax % stays a screener-input value (no Tax/PBT row exists on this sheet).
        # D&A rate on PP&E cross-references the BS Net Fixed Assets row (R_PPE_BS=10).
        BSN_IS = "'5_Balance_Sheet'"
        R_PPE_BS = 10                       # Net Fixed Assets row on the BS sheet

        htax = h_is.get("tax_pct") or []    # screener effective tax % (e.g. 25.4)
        def _g_val(a, i):
            return (a[i] if (i < len(a) and a[i]) else None)

        for i in range(n):
            col_i = get_column_letter(3 + i)    # this column (current actual year)
            col_p = get_column_letter(3 + i - 1)  # prior column (previous actual year)

            # Revenue growth %: blank for first column (no prior year)
            if i > 0:
                f = ws.cell(row=R_G, column=3 + i,
                            value=f"={col_i}{R_REV}/{col_p}{R_REV}-1")
                f.number_format = F.FMT_PERCENT; f.font = F.FONT_INPUT

            # COGS % sales
            f = ws.cell(row=R_COGSP, column=3 + i,
                        value=f"={col_i}{R_COGS}/{col_i}{R_REV}")
            f.number_format = F.FMT_PERCENT; f.font = F.FONT_INPUT

            # EBIT margin %
            f = ws.cell(row=R_MG, column=3 + i,
                        value=f"={col_i}{R_EBIT}/{col_i}{R_REV}")
            f.number_format = F.FMT_PERCENT; f.font = F.FONT_INPUT

            # Tax % — keep as screener-input value (no Tax/PBT rows on IS sheet)
            tx = _g_val(htax, i)
            if tx:
                c = ws.cell(row=R_TAX, column=3 + i, value=tx / 100.0)
                c.number_format = F.FMT_PERCENT; c.font = F.FONT_INPUT

            # NOPAT = EBIT × (1 − Tax %) — formula refs same-column cells
            f = ws.cell(row=R_NOP, column=3 + i,
                        value=f"={col_i}{R_EBIT}*(1-{col_i}{R_TAX})")
            f.number_format = F.FMT_NUMBER; f.font = F.FONT_INPUT

            # CapEx % sales
            f = ws.cell(row=R_CXP, column=3 + i,
                        value=f"={col_i}{R_CX}/{col_i}{R_REV}")
            f.number_format = F.FMT_PERCENT; f.font = F.FONT_INPUT

            # D&A rate on PP&E: D&A[i] / Net Fixed Assets[i-1] — cross-sheet BS ref.
            # First actual year has no prior-year PPE column on BS → leave blank.
            if i > 0:
                f = ws.cell(row=R_DAP, column=3 + i,
                            value=f"={col_i}{R_DA}/{BSN_IS}!{col_p}{R_PPE_BS}")
                f.number_format = F.FMT_PERCENT; f.font = F.FONT_INPUT

        # Projected Net Profit (formula). Levered: NOPAT - Interest * (1 - Tax)
        DEBTN = "'7_Debt_Schedule'"
        R_INT_DEBT = 9
        for j in range(ny):
            cc = pc0 + j
            ws.cell(row=R_INT, column=cc, value=f"={DEBTN}!{L(cc)}{R_INT_DEBT}").number_format = F.FMT_NUMBER
            ws.cell(row=R_NP, column=cc, value=f"={L(cc)}{R_NOP}-{L(cc)}{R_INT}*(1-{L(cc)}{R_TAX})").number_format = F.FMT_NUMBER

        # Expose cell coordinates for the FCF/BS/CF sheets
        self._is_ref = {"pc0": pc0, "ny": ny, "R_NOP": R_NOP, "R_DA": R_DA,
                        "R_CX": R_CX, "R_EBIT": R_EBIT, "R_NP": R_NP, "R_REV": R_REV, "R_COGS": R_COGS}

        # Per-year projection drivers from the engine (editable blue inputs)
        prev = hist_rev[-1] if hist_rev else (proj["revenue"][0])
        gr = []
        for v in proj["revenue"]:
            gr.append((v / prev - 1.0) if prev else 0.0); prev = v
        cogsp = [(c / r if r else 0.0) for c, r in zip(proj["cogs"], proj["revenue"])]
        mg = [(e / r if r else 0.0) for e, r in zip(proj["ebit"], proj["revenue"])]
        cxp = [(c / r if r else 0.0) for c, r in zip(proj["capex"], proj["revenue"])]
        tax = au.get("tax_rate", 0.25)
        dep_rate = au.get("da_rate_on_block", 0.08)   # D&A as % of opening net block (PP&E schedule)
        BSN = "'5_Balance_Sheet'"
        R_PPE_BS = 10                                  # Net Fixed Assets row on the BS sheet
        last_ppe_is = hpe[-1] if hpe else 0.0          # last actual net block

        for j in range(ny):
            cc = pc0 + j
            Lc, Lp = L(cc), L(cc - 1)
            for row, val in [(R_G, gr[j]), (R_COGSP, cogsp[j]), (R_MG, mg[j]), (R_TAX, tax), (R_CXP, cxp[j]), (R_DAP, dep_rate)]:
                c = ws.cell(row=row, column=cc, value=val)
                c.number_format = F.FMT_PERCENT; c.font = F.FONT_INPUT  # blue editable driver
            ws.cell(row=R_REV, column=cc, value=f"={Lp}{R_REV}*(1+{Lc}{R_G})").number_format = F.FMT_NUMBER
            ws.cell(row=R_COGS, column=cc, value=f"={Lc}{R_REV}*{Lc}{R_COGSP}").number_format = F.FMT_NUMBER
            ws.cell(row=R_EBIT, column=cc, value=f"={Lc}{R_REV}*{Lc}{R_MG}").number_format = F.FMT_NUMBER
            ws.cell(row=R_NOP, column=cc, value=f"={Lc}{R_EBIT}*(1-{Lc}{R_TAX})").number_format = F.FMT_NUMBER
            ws.cell(row=R_CX, column=cc, value=f"={Lc}{R_REV}*{Lc}{R_CXP}").number_format = F.FMT_NUMBER
            # D&A from the PP&E schedule: opening (prior-year) net block × dep rate.
            # First projection year's opening block = last actual net block (static).
            ppe_open = f"{BSN}!{Lp}{R_PPE_BS}" if j else repr(last_ppe_is)
            ws.cell(row=R_DA, column=cc, value=f"={ppe_open}*{Lc}{R_DAP}").number_format = F.FMT_NUMBER

    def render_bs(self):
        """Integrated, balancing Balance Sheet. Projections are LIVE formulas that
        link to the Income Statement (Revenue, CapEx, D&A, Net Profit): AR/Inv/AP
        from DSO/DIO/DPO × revenue, PP&E rolls (prior + CapEx − D&A), Equity rolls
        (prior + Net Profit × (1−payout)), Cash rolls (prior + NP + D&A − ΔNWC − CapEx
        − Dividends), debt is held flat and a constant 'other net assets' plug makes
        year-0 tie. Net income is pre-financing (≈ NOPAT) so there is NO circular
        reference — opens clean — yet the Balance Check is 0 every year. Edit any
        driver and it re-balances."""
        from openpyxl.utils import get_column_letter as L
        F = Formats
        ws = self._create_sheet("5_Balance_Sheet")
        n = self._av_header(ws, f"Balance Sheet ({self._unit}) — integrated / balancing")
        ref = self._is_ref
        R_REV, R_COGS, R_CX, R_DA, R_NP = ref["R_REV"], ref["R_COGS"], ref["R_CX"], ref["R_DA"], ref["R_NP"]
        ny = ref["ny"]
        ISN = "'4_Income_Statement'"
        hb = self.results.get("hist", {}).get("bs", {})
        au = self.results["proj"].get("assumptions_used", {})

        def hl(key):
            s = hb.get(key) or []
            return (s[-1] if s and s[-1] else 0.0)
        last_cash, last_ppe = hl("cash_equivalents"), hl("fixed_assets")
        last_ar, last_inv, last_ap = hl("trade_receivables"), hl("inventories"), hl("trade_payables")
        last_debt = hl("borrowings") + hl("lease_liabilities")
        last_eq = hl("equity_capital") + hl("reserves")
        last_nwc = last_ar + last_inv - last_ap
        last_owc = 0.0
        other = (last_ap + last_debt + last_eq) - (last_cash + last_ar + last_inv + last_ppe)
        dso, dio, dpo = au.get("dso_days", 45.0), au.get("dio_days", 30.0), au.get("dpo_days", 45.0)
        # Financials freeze operating working capital flat at the historical anchor
        frozen_wc = self.results["proj"].get("freeze_working_capital", False)

        # Working Capital Days model
        wc_days = au.get("working_capital_days", None)
        his = self.results.get("hist", {}).get("is", {})
        _sales = his.get("sales") or his.get("revenue") or []
        last_sales = (_sales[-1] if _sales and _sales[-1] else 0.0)
        use_wc_days = (wc_days is not None) and not frozen_wc
        if use_wc_days:
            nwc_net_0 = (wc_days / 365.0) * last_sales
            last_nwc = nwc_net_0
            last_owc = nwc_net_0 - (last_ar + last_inv - last_ap)
            other = (last_ap + last_debt + last_eq) - (last_cash + last_ar + last_inv + last_owc + last_ppe)

        # Dividend payout
        div_payout = au.get("dividend_payout_ratio", 0.0) or 0.0

        # Row layout
        R_DSO, R_DIO, R_DPO = 3, 4, 5
        R_CASH, R_AR, R_INV, R_OWC, R_PPE, R_OTH, R_TA = 6, 7, 8, 9, 10, 11, 12
        R_AP, R_DEBT, R_EQ = 13, 14, 15
        R_PAYP, R_DIV = 16, 17          # new: payout % driver + dividends outflow
        R_TLE, R_NWC, R_BC = 18, 19, 20

        for rr, lab in [(R_DSO, "DSO (days)"), (R_DIO, "DIO (days)"), (R_DPO, "DPO (days)"),
                        (R_CASH, "Cash"), (R_AR, "Trade Receivables"), (R_INV, "Inventories"),
                        (R_OWC, "Other net working capital"), (R_PPE, "Net Fixed Assets"),
                        (R_OTH, "Other net assets (plug)"), (R_TA, "Total Assets"),
                        (R_AP, "Trade Payables"), (R_DEBT, "Debt"), (R_EQ, "Equity"),
                        (R_PAYP, "Dividend payout %"), (R_DIV, "Dividends"),
                        (R_TLE, "Total Liab + Equity"), (R_NWC, "Net Working Capital"),
                        (R_BC, "Balance Check")]:
            ws.cell(row=rr, column=2, value=lab).font = F.FONT_BOLD

        # ── Historical input rows (blue values) ──────────────────────────────
        def hist(row, key):
            for i, v in enumerate((hb.get(key) or [])[-n:] if n else []):
                if v:
                    c = ws.cell(row=row, column=3 + i, value=v)
                    c.number_format = F.FMT_NUMBER; c.font = F.FONT_INPUT
        hist(R_CASH, "cash_equivalents"); hist(R_AR, "trade_receivables")
        hist(R_INV, "inventories"); hist(R_PPE, "fixed_assets"); hist(R_AP, "trade_payables")

        # Historical Equity (capital + reserves)
        h_eq_c = (hb.get("equity_capital") or [])[-n:] if n else []
        h_res  = (hb.get("reserves")       or [])[-n:] if n else []
        for i in range(n):
            eq_v = (h_eq_c[i] if i < len(h_eq_c) else 0.0) + (h_res[i] if i < len(h_res) else 0.0)
            if eq_v:
                c = ws.cell(row=R_EQ, column=3 + i, value=eq_v)
                c.number_format = F.FMT_NUMBER; c.font = F.FONT_INPUT

        # Historical Debt
        h_bor = (hb.get("borrowings")        or [])[-n:] if n else []
        h_lea = (hb.get("lease_liabilities") or [])[-n:] if n else []
        for i in range(n):
            debt_v = (h_bor[i] if i < len(h_bor) else 0.0) + (h_lea[i] if i < len(h_lea) else 0.0)
            if debt_v:
                c = ws.cell(row=R_DEBT, column=3 + i, value=debt_v)
                c.number_format = F.FMT_NUMBER; c.font = F.FONT_INPUT

        # Historical Dividend payout % (screener value)
        h_pay = (self.results.get("hist", {}).get("is", {}).get("dividend_payout_pct") or [])[-n:] if n else []
        for i in range(n):
            pv = h_pay[i] if i < len(h_pay) else 0.0
            if pv:
                c = ws.cell(row=R_PAYP, column=3 + i, value=pv / 100.0)
                c.number_format = F.FMT_PERCENT; c.font = F.FONT_INPUT

        # ── Historical DERIVED rows — LIVE Excel formulas (same pattern as projection) ──
        # DSO = AR / IS.Revenue × 365
        # DIO = INV / IS.COGS × 365
        # DPO = AP / IS.COGS × 365
        # OWC = NWC − (AR+INV−AP)       (or 0 when no wc_days in actuals)
        # OTH = stays a Python-computed plug (it IS the balancing item)
        # TA  = CASH+AR+INV+OWC+PPE+OTH
        # TLE = AP+DEBT+EQ
        # NWC = AR+INV−AP+OWC
        # BC  = TA−TLE
        # Dividends (historical) = NI (from IS) × payout %
        _ratios_h = self.results.get("hist", {}).get("ratios", {}) or {}
        _rev_h = his.get("sales") or his.get("revenue") or []

        def _g(a, i):
            return (a[i] if (i < len(a) and a[i] is not None) else 0.0)
        def _ln(a):
            return (a or [])[-n:] if n else []

        _ar_h  = _ln(hb.get("trade_receivables"))
        _inv_h = _ln(hb.get("inventories"))
        _ap_h  = _ln(hb.get("trade_payables"))
        _bor_h = _ln(hb.get("borrowings"))
        _lea_h = _ln(hb.get("lease_liabilities"))
        _eqc_h = _ln(hb.get("equity_capital"))
        _res_h = _ln(hb.get("reserves"))
        _wcd_h = _ln(_ratios_h.get("working_capital_days"))
        _revh  = _ln(_rev_h)
        _ppe_h = _ln(hb.get("fixed_assets"))
        _cash_h = _ln(hb.get("cash_equivalents"))
        _last_nwc_hist = last_nwc

        for i in range(n):
            col = L(3 + i)
            ar_v   = _g(_ar_h,  i)
            inv_v  = _g(_inv_h, i)
            ap_v   = _g(_ap_h,  i)
            cash_v = _g(_cash_h, i)
            ppe_v  = _g(_ppe_h,  i)
            debt_v = _g(_bor_h, i) + _g(_lea_h, i)
            eq_v   = _g(_eqc_h, i) + _g(_res_h, i)
            rev_v  = _g(_revh,  i)
            wcd_v  = _g(_wcd_h, i)

            # OWC and plug (Python values — OTH is the plug, not a formula)
            trade = ar_v + inv_v - ap_v
            nwc_v = (wcd_v / 365.0) * rev_v if (wcd_v and rev_v) else trade
            owc_v = nwc_v - trade
            plug  = (ap_v + debt_v + eq_v) - (cash_v + ar_v + inv_v + owc_v + ppe_v)

            def _fw(row, val, fmt=F.FMT_NUMBER):
                c = ws.cell(row=row, column=3 + i, value=val)
                c.number_format = fmt; c.font = F.FONT_INPUT

            # DSO/DIO/DPO: formula when balance is non-zero; else blank
            if ar_v and rev_v:
                f = ws.cell(row=R_DSO, column=3 + i,
                            value=f"={col}{R_AR}/{ISN}!{col}{R_REV}*365")
                f.number_format = '0.0'; f.font = F.FONT_INPUT
            if inv_v and rev_v:
                f = ws.cell(row=R_DIO, column=3 + i,
                            value=f"={col}{R_INV}/{ISN}!{col}{R_COGS}*365")
                f.number_format = '0.0'; f.font = F.FONT_INPUT
            if ap_v and rev_v:
                f = ws.cell(row=R_DPO, column=3 + i,
                            value=f"={col}{R_AP}/{ISN}!{col}{R_COGS}*365")
                f.number_format = '0.0'; f.font = F.FONT_INPUT

            # OWC: formula
            ws.cell(row=R_OWC, column=3 + i,
                    value=f"={col}{R_NWC}-({col}{R_AR}+{col}{R_INV}-{col}{R_AP})"
                    ).number_format = F.FMT_NUMBER

            # Plug (OTH): Python value — it IS the balancing residual
            _fw(R_OTH, plug)

            # Total Assets: formula
            ws.cell(row=R_TA, column=3 + i,
                    value=f"={col}{R_CASH}+{col}{R_AR}+{col}{R_INV}+{col}{R_OWC}+{col}{R_PPE}+{col}{R_OTH}"
                    ).number_format = F.FMT_NUMBER

            # Total Liab + Equity: formula
            ws.cell(row=R_TLE, column=3 + i,
                    value=f"={col}{R_AP}+{col}{R_DEBT}+{col}{R_EQ}"
                    ).number_format = F.FMT_NUMBER

            # NWC: formula
            if wcd_v and rev_v:
                ws.cell(row=R_NWC, column=3 + i,
                        value=f"={ISN}!{col}{R_REV}*{wcd_v}/365"
                        ).number_format = F.FMT_NUMBER
            else:
                ws.cell(row=R_NWC, column=3 + i,
                        value=f"={col}{R_AR}+{col}{R_INV}-{col}{R_AP}+{col}{R_OWC}"
                        ).number_format = F.FMT_NUMBER

            # Balance Check: formula
            ws.cell(row=R_BC, column=3 + i,
                    value=f"={col}{R_TA}-{col}{R_TLE}"
                    ).number_format = F.FMT_NUMBER

            # Dividends (historical): IS NP × payout %
            ws.cell(row=R_DIV, column=3 + i,
                    value=f"={ISN}!{col}{R_NP}*{col}{R_PAYP}"
                    ).number_format = F.FMT_NUMBER

            if i == n - 1:
                _last_nwc_hist = nwc_v

        # Stash NWC anchor for the Cash Flow sheet's live ΔNWC formula.
        self._bs_nwc = {"row": R_NWC, "anchor": last_nwc, "pc0": 3 + n}

        # ── Projection columns ────────────────────────────────────────────────
        for j in range(ny):
            cc = 3 + n + j
            col, colp = L(cc), L(cc - 1)
            if frozen_wc:
                for rr, val in [(R_AR, last_ar), (R_INV, last_inv), (R_AP, last_ap), (R_OWC, last_owc)]:
                    c = ws.cell(row=rr, column=cc, value=val)
                    c.number_format = F.FMT_NUMBER; c.font = F.FONT_INPUT
                ws.cell(row=R_NWC, column=cc, value=repr(last_nwc)).number_format = F.FMT_NUMBER
                dnwc = "0"
            elif use_wc_days:
                for rr, val in [(R_DSO, dso), (R_DIO, dio), (R_DPO, dpo)]:
                    c = ws.cell(row=rr, column=cc, value=val); c.number_format = '0.0'; c.font = F.FONT_INPUT
                ws.cell(row=R_AR, column=cc, value=f"={ISN}!{col}{R_REV}*{col}{R_DSO}/365").number_format = F.FMT_NUMBER
                ws.cell(row=R_INV, column=cc, value=f"={ISN}!{col}{R_COGS}*{col}{R_DIO}/365").number_format = F.FMT_NUMBER
                ws.cell(row=R_AP, column=cc, value=f"={ISN}!{col}{R_COGS}*{col}{R_DPO}/365").number_format = F.FMT_NUMBER
                ws.cell(row=R_NWC, column=cc, value=f"={ISN}!{col}{R_REV}*{repr(wc_days)}/365").number_format = F.FMT_NUMBER
                ws.cell(row=R_OWC, column=cc, value=f"={col}{R_NWC}-({col}{R_AR}+{col}{R_INV}-{col}{R_AP})").number_format = F.FMT_NUMBER
                dnwc = f"({col}{R_NWC}-{colp}{R_NWC})" if j else f"({col}{R_NWC}-{repr(last_nwc)})"
            else:
                for rr, val in [(R_DSO, dso), (R_DIO, dio), (R_DPO, dpo)]:
                    c = ws.cell(row=rr, column=cc, value=val); c.number_format = '0.0'; c.font = F.FONT_INPUT
                ws.cell(row=R_AR, column=cc, value=f"={ISN}!{col}{R_REV}*{col}{R_DSO}/365").number_format = F.FMT_NUMBER
                ws.cell(row=R_INV, column=cc, value=f"={ISN}!{col}{R_COGS}*{col}{R_DIO}/365").number_format = F.FMT_NUMBER
                ws.cell(row=R_AP, column=cc, value=f"={ISN}!{col}{R_COGS}*{col}{R_DPO}/365").number_format = F.FMT_NUMBER
                ws.cell(row=R_OWC, column=cc, value=0.0).number_format = F.FMT_NUMBER
                ws.cell(row=R_NWC, column=cc, value=f"={col}{R_AR}+{col}{R_INV}-{col}{R_AP}+{col}{R_OWC}").number_format = F.FMT_NUMBER
                dnwc = f"({col}{R_NWC}-{colp}{R_NWC})" if j else f"({col}{R_NWC}-{repr(last_nwc)})"

            ppe_prev = f"{colp}{R_PPE}" if j else repr(last_ppe)
            ws.cell(row=R_PPE, column=cc, value=f"={ppe_prev}+{ISN}!{col}{R_CX}-{ISN}!{col}{R_DA}").number_format = F.FMT_NUMBER

            # Dividend payout % driver row (projected — blue editable input)
            c_pay = ws.cell(row=R_PAYP, column=cc, value=div_payout)
            c_pay.number_format = F.FMT_PERCENT; c_pay.font = F.FONT_INPUT

            # Dividends = NP × payout %
            ws.cell(row=R_DIV, column=cc,
                    value=f"={ISN}!{col}{R_NP}*{col}{R_PAYP}"
                    ).number_format = F.FMT_NUMBER

            eq_prev = f"{colp}{R_EQ}" if j else repr(last_eq)
            # Equity roll: prior + NP × (1 − payout)
            ws.cell(row=R_EQ, column=cc,
                    value=f"={eq_prev}+{ISN}!{col}{R_NP}*(1-{col}{R_PAYP})"
                    ).number_format = F.FMT_NUMBER

            DEBTN = "'7_Debt_Schedule'"
            R_CLOSE_DEBT = 7
            cash_prev = f"{colp}{R_CASH}" if j else repr(last_cash)
            # Cash roll: prior + NP + D&A − ΔNWC − CapEx − Dividends + (Debt_t - Debt_{t-1})
            debt_diff = f"({DEBTN}!{col}{R_CLOSE_DEBT}-{DEBTN}!{colp}{R_CLOSE_DEBT})" if j else f"({DEBTN}!{col}{R_CLOSE_DEBT}-{repr(last_debt)})"
            ws.cell(row=R_CASH, column=cc,
                    value=f"={cash_prev}+{ISN}!{col}{R_NP}+{ISN}!{col}{R_DA}-{dnwc}-{ISN}!{col}{R_CX}-{col}{R_DIV}+{debt_diff}"
                    ).number_format = F.FMT_NUMBER

            ws.cell(row=R_DEBT, column=cc, value=f"={DEBTN}!{col}{R_CLOSE_DEBT}").number_format = F.FMT_NUMBER
            ocell = ws.cell(row=R_OTH, column=cc, value=other); ocell.number_format = F.FMT_NUMBER
            ws.cell(row=R_TA, column=cc, value=f"={col}{R_CASH}+{col}{R_AR}+{col}{R_INV}+{col}{R_OWC}+{col}{R_PPE}+{col}{R_OTH}").number_format = F.FMT_NUMBER
            ws.cell(row=R_TLE, column=cc, value=f"={col}{R_AP}+{col}{R_DEBT}+{col}{R_EQ}").number_format = F.FMT_NUMBER
            ws.cell(row=R_BC, column=cc, value=f"={col}{R_TA}-{col}{R_TLE}").number_format = F.FMT_NUMBER

    def render_cf(self):
        """Cash Flow linked to the Income Statement: Net Income, D&A and CapEx are
        formulas pulling from the IS; CFO and FCF are formulas. CapEx is shown as a
        consistent outflow (negative) in both actuals and projections."""
        from openpyxl.utils import get_column_letter as L
        F = Formats
        ws = self._create_sheet("6_Cash_Flow")
        n = self._av_header(ws, f"Cash Flow ({self._unit}) — linked to Income Statement")
        p = self.results["proj"]
        ref = getattr(self, "_is_ref", {})
        R_NP, R_DA, R_CX = ref.get("R_NP", 13), ref.get("R_DA", 12), ref.get("R_CX", 10)
        ny = len(p["years"])
        ISN = "'4_Income_Statement'"
        h_is = self.results.get("hist", {}).get("is", {})
        hc = self.results.get("hist", {}).get("cf", {})
        hist_ni = h_is.get("net_profit") or []
        hist_da = h_is.get("depreciation") or []
        hist_cfo = hc.get("cfo") or []
        fap = hc.get("fixed_assets_purchased") or []
        hist_capex = [-(abs(x)) if x else 0.0 for x in fap]            # outflow (negative)
        hist_fcf = [(c or 0.0) - abs(x or 0.0) for c, x in zip(hist_cfo, fap)]

        for rr, lab in [(3, "Net Income"), (4, "Add: Depreciation"), (5, "Less: Change in NWC"),
                        (6, "Cash from Operations"), (7, "Less: CapEx"), (8, "Free Cash Flow"),
                        (9, "Proceeds from Borrowings"), (10, "Repayment of Borrowings"),
                        (11, "Net Change in Cash")]:
            ws.cell(row=rr, column=2, value=lab).font = F.FONT_BOLD

        def hist(row, vals):
            for i, v in enumerate(vals[-n:] if n else []):
                if v:
                    c = ws.cell(row=row, column=3 + i, value=v)
                    c.number_format = F.FMT_NUMBER; c.font = F.FONT_INPUT
        hist_proc = hc.get("proceeds_from_borrowings") or []
        hist_rep = hc.get("repayment_of_borrowings") or []
        hist_rep = [-abs(x) if x else 0.0 for x in hist_rep] # outflow
        hist_net_cash = [(c or 0.0) + (f or 0.0) + (p or 0.0) + (r or 0.0) for c, f, p, r in zip(hist_cfo, hist_capex, hist_proc, hist_rep)]
        
        hist(3, hist_ni); hist(4, hist_da); hist(6, hist_cfo); hist(7, hist_capex); hist(8, hist_fcf)
        hist(9, hist_proc); hist(10, hist_rep); hist(11, hist_net_cash)

        # ΔNWC is no longer hardcoded — it is computed live as −(NWC_t − NWC_{t-1})
        # off the Balance Sheet's Net Working Capital row (which is driven by Working
        # Capital Days). First projection year anchors on the engine's Year-0 NWC.
        BSN = "'5_Balance_Sheet'"
        bsref = getattr(self, "_bs_nwc", {})
        R_NWC_BS = bsref.get("row", 19)  # NWC row is 19 after dividend rows added
        nwc_anchor = bsref.get("anchor", 0.0)

        # Historical Change in NWC (row 5): −Δ of the historical NWC on the BS sheet.
        for i in range(1, n):
            cprev, ccur = L(3 + i - 1), L(3 + i)
            ch = ws.cell(row=5, column=3 + i, value=f"=-({BSN}!{ccur}{R_NWC_BS}-{BSN}!{cprev}{R_NWC_BS})")
            ch.number_format = F.FMT_NUMBER; ch.font = F.FONT_LINK

        DEBTN = "'7_Debt_Schedule'"
        R_PROC_DEBT = 5
        R_REP_DEBT = 6
        for j in range(ny):
            cc = 3 + n + j            # projection column (aligns to the IS/BS projection column)
            col, colp = L(cc), L(cc - 1)
            prevref = repr(nwc_anchor) if j == 0 else f"{BSN}!{colp}{R_NWC_BS}"
            c5 = ws.cell(row=5, column=cc, value=f"=-({BSN}!{col}{R_NWC_BS}-{prevref})")
            c5.number_format = F.FMT_NUMBER; c5.font = F.FONT_LINK
            l1 = ws.cell(row=3, column=cc, value=f"={ISN}!{col}{R_NP}"); l1.number_format = F.FMT_NUMBER; l1.font = F.FONT_LINK
            l2 = ws.cell(row=4, column=cc, value=f"={ISN}!{col}{R_DA}"); l2.number_format = F.FMT_NUMBER; l2.font = F.FONT_LINK
            ws.cell(row=6, column=cc, value=f"={col}3+{col}4+{col}5").number_format = F.FMT_NUMBER
            l3 = ws.cell(row=7, column=cc, value=f"=-{ISN}!{col}{R_CX}"); l3.number_format = F.FMT_NUMBER; l3.font = F.FONT_LINK
            ws.cell(row=8, column=cc, value=f"={col}6+{col}7").number_format = F.FMT_NUMBER
            ws.cell(row=9, column=cc, value=f"={DEBTN}!{col}{R_PROC_DEBT}").number_format = F.FMT_NUMBER
            ws.cell(row=10, column=cc, value=f"={DEBTN}!{col}{R_REP_DEBT}").number_format = F.FMT_NUMBER
            ws.cell(row=11, column=cc, value=f"={col}8+{col}9+{col}10").number_format = F.FMT_NUMBER

    def render_debt(self):
        F = Formats
        from openpyxl.utils import get_column_letter as L
        ws = self._create_sheet("7_Debt_Schedule")
        n = self._av_header(ws, f"Debt Schedule ({self._unit}) — Historical & Projected")
        p = self.results["proj"]
        h = self.results.get("hist", {})
        hbs = h.get("bs", {})
        hcf = h.get("cf", {})
        his = h.get("is", {})
        ny = len(p["years"])
        
        R_RATIO = 3
        R_OPEN = 4
        R_PROC = 5
        R_REP = 6
        R_CLOSE = 7
        R_RATE = 8
        R_INT = 9
        
        for rr, lab in [(R_RATIO, "Debt / EBITDA ratio target"), (R_OPEN, "Opening Debt"), 
                        (R_PROC, "Proceeds from Borrowings"), (R_REP, "Repayment of Borrowings"), 
                        (R_CLOSE, "Closing Debt"), (R_RATE, "Effective Interest Rate %"), 
                        (R_INT, "Interest Expense")]:
            ws.cell(row=rr, column=2, value=lab).font = F.FONT_BOLD
            
        def _ln(a): return (a or [])[-n:] if n else []
        def _g(a, i): return (a[i] if (i < len(a) and a[i] is not None) else 0.0)
        
        h_bor = _ln(hbs.get("borrowings"))
        h_lea = _ln(hbs.get("lease_liabilities"))
        h_proc = _ln(hcf.get("proceeds_from_borrowings"))
        h_rep = _ln(hcf.get("repayment_of_borrowings"))
        h_int = _ln(his.get("interest"))
        
        for i in range(n):
            col = L(3 + i)
            colp = L(3 + i - 1)
            
            c_close = ws.cell(row=R_CLOSE, column=3 + i, value=_g(h_bor, i) + _g(h_lea, i))
            c_close.number_format = F.FMT_NUMBER; c_close.font = F.FONT_INPUT
            
            if _g(h_proc, i):
                c_proc = ws.cell(row=R_PROC, column=3 + i, value=_g(h_proc, i))
                c_proc.number_format = F.FMT_NUMBER; c_proc.font = F.FONT_INPUT
                
            if _g(h_rep, i):
                c_rep = ws.cell(row=R_REP, column=3 + i, value=-abs(_g(h_rep, i)))
                c_rep.number_format = F.FMT_NUMBER; c_rep.font = F.FONT_INPUT
                
            if _g(h_int, i):
                c_int = ws.cell(row=R_INT, column=3 + i, value=_g(h_int, i))
                c_int.number_format = F.FMT_NUMBER; c_int.font = F.FONT_INPUT
                
            if i > 0:
                ws.cell(row=R_OPEN, column=3 + i, value=f"={colp}{R_CLOSE}").number_format = F.FMT_NUMBER
                if _g(h_bor, i-1) + _g(h_lea, i-1) > 0:
                    ws.cell(row=R_RATE, column=3 + i, value=f"={col}{R_INT}/{colp}{R_CLOSE}").number_format = F.FMT_PERCENT
                
        ISN = "'4_Income_Statement'"
        ref = getattr(self, "_is_ref", {})
        R_EBIT = ref.get("R_EBIT", 8); R_DA = ref.get("R_DA", 14)
        
        au = p.get("assumptions_used", {})
        target_ratio = au.get("debt_ebitda_ratio")
        eff_rate = au.get("effective_rate", 0.08)
        
        for j in range(ny):
            cc = 3 + n + j
            col = L(cc)
            colp = L(cc - 1)
            
            if target_ratio is not None:
                cr = ws.cell(row=R_RATIO, column=cc, value=target_ratio)
                cr.number_format = F.FMT_NUMBER; cr.font = F.FONT_INPUT
            
            crate = ws.cell(row=R_RATE, column=cc, value=eff_rate)
            crate.number_format = F.FMT_PERCENT; crate.font = F.FONT_INPUT
            
            ws.cell(row=R_OPEN, column=cc, value=f"={colp}{R_CLOSE}").number_format = F.FMT_NUMBER
            
            if target_ratio is not None:
                ebitda_ref = f"({ISN}!{col}{R_EBIT}+{ISN}!{col}{R_DA})"
                ws.cell(row=R_CLOSE, column=cc, value=f"=IF({ebitda_ref}>0, {col}{R_RATIO}*{ebitda_ref}, {colp}{R_CLOSE})").number_format = F.FMT_NUMBER
                ws.cell(row=R_CLOSE, column=cc).font = F.FONT_BOLD
            else:
                ws.cell(row=R_CLOSE, column=cc, value=f"={colp}{R_CLOSE}").number_format = F.FMT_NUMBER
                ws.cell(row=R_CLOSE, column=cc).font = F.FONT_BOLD
                
            ws.cell(row=R_PROC, column=cc, value=f"=MAX(0, {col}{R_CLOSE}-{col}{R_OPEN})").number_format = F.FMT_NUMBER
            ws.cell(row=R_REP, column=cc, value=f"=MIN(0, {col}{R_CLOSE}-{col}{R_OPEN})").number_format = F.FMT_NUMBER
            ws.cell(row=R_INT, column=cc, value=f"={col}{R_RATE}*{col}{R_OPEN}").number_format = F.FMT_NUMBER

    def render_fcf(self):
        """Live DCF: UFCF = NOPAT + D&A − CapEx − ΔNWC (NOPAT/D&A/CapEx linked from the
        Income Statement), discounted at the WACC (linked from 9_WACC) with a Gordon
        terminal at g (linked from 10_Terminal). Editing any IS driver, the WACC, or g
        recomputes the Enterprise Value here."""
        from openpyxl.utils import get_column_letter as L
        F = Formats
        ws = self._create_sheet("8_FCF_DCF")
        p = self.results["proj"]
        ref = getattr(self, "_is_ref", None)
        if not ref:
            # Fallback to plain values if the IS layout wasn't captured
            self._proj_header(ws, f"Unlevered FCF & Discounting ({self._unit})")
            self._row(ws, 3, "Unlevered FCF", p["ufcf"], bold=True)
            return
        pc0, ny = ref["pc0"], ref["ny"]
        R_NOP, R_DA, R_CX = ref["R_NOP"], ref["R_DA"], ref["R_CX"]
        ISN = "'4_Income_Statement'"

        c = ws.cell(row=2, column=2, value=f"Unlevered FCF & DCF ({self._unit})")
        c.font = F.FONT_HEADER; c.fill = F.FILL_HEADER
        for j, y in enumerate(p["years"]):
            hc = ws.cell(row=2, column=3 + j, value=self._fy_label(y, "E"))
            hc.font = F.FONT_HEADER; hc.fill = F.FILL_HEADER; hc.alignment = F.ALIGN_CENTER

        for rr, lab in [(3, "Unlevered FCF"), (4, "Change in NWC"), (5, "Discount period"),
                        (6, "Discount factor"), (7, "PV of UFCF")]:
            ws.cell(row=rr, column=2, value=lab).font = F.FONT_BOLD

        nwc = p.get("nwc_change", [0.0] * ny)
        for j in range(ny):
            cc = pc0 + j          # matching IS column
            isL = L(cc)
            lc = L(3 + j)         # local column
            # ΔNWC (editable value)
            nv = ws.cell(row=4, column=3 + j, value=(nwc[j] if j < len(nwc) else 0.0))
            nv.number_format = F.FMT_NUMBER; nv.font = F.FONT_INPUT
            # UFCF = NOPAT + D&A − CapEx − ΔNWC
            ws.cell(row=3, column=3 + j,
                    value=f"={ISN}!{isL}{R_NOP}+{ISN}!{isL}{R_DA}-{ISN}!{isL}{R_CX}-{lc}4"
                    ).number_format = F.FMT_NUMBER
            # discount period (mid-year), factor, PV
            pcell = ws.cell(row=5, column=3 + j, value=0.5 + j); pcell.number_format = '0.0'
            ws.cell(row=6, column=3 + j, value=f"=1/(1+$C$9)^{lc}5").number_format = '0.000'
            ws.cell(row=7, column=3 + j, value=f"={lc}3*{lc}6").number_format = F.FMT_NUMBER

        last = L(3 + ny - 1)            # last UFCF local column
        isLast = L(pc0 + ny - 1)        # last projection column on the IS sheet
        R_EBIT = ref["R_EBIT"]
        # Summary block (col C). WACC/g/exit-multiple linked from their detailed sheets.
        # Terminal value matches the engine: average of Gordon and exit-multiple TV.
        ws.cell(row=9, column=2, value="WACC (from 9_WACC)").font = F.FONT_BOLD
        w9 = ws.cell(row=9, column=3, value="='9_WACC'!$C$18"); w9.font = F.FONT_LINK; w9.number_format = F.FMT_PERCENT
        ws.cell(row=10, column=2, value="Terminal growth g (from 10_Terminal)").font = F.FONT_BOLD
        g10 = ws.cell(row=10, column=3, value="='10_Terminal'!$C$3"); g10.font = F.FONT_LINK; g10.number_format = F.FMT_PERCENT
        ws.cell(row=11, column=2, value="Exit EV/EBITDA (from 10_Terminal)").font = F.FONT_BOLD
        x11 = ws.cell(row=11, column=3, value="='10_Terminal'!$C$4"); x11.font = F.FONT_LINK; x11.number_format = F.FMT_MULTIPLE
        ws.cell(row=12, column=2, value="Sum PV of UFCF").font = F.FONT_BOLD
        ws.cell(row=12, column=3, value=f"=SUM(C7:{last}7)").number_format = F.FMT_NUMBER
        ws.cell(row=13, column=2, value="Gordon TV").font = F.FONT_BOLD
        ws.cell(row=13, column=3, value=f"={last}3*(1+$C$10)/($C$9-$C$10)").number_format = F.FMT_NUMBER
        ws.cell(row=14, column=2, value="Exit-multiple TV").font = F.FONT_BOLD
        ws.cell(row=14, column=3, value=f"=({ISN}!{isLast}{R_EBIT}+{ISN}!{isLast}{R_DA})*$C$11").number_format = F.FMT_NUMBER
        ws.cell(row=15, column=2, value="Average TV (used)").font = F.FONT_BOLD
        ws.cell(row=15, column=3, value="=(C13+C14)/2").number_format = F.FMT_NUMBER
        ws.cell(row=16, column=2, value="PV of Terminal Value").font = F.FONT_BOLD
        ws.cell(row=16, column=3, value=f"=C15/(1+$C$9)^{ny}").number_format = F.FMT_NUMBER
        ws.cell(row=17, column=2, value="Enterprise Value").font = F.FONT_BOLD
        ws.cell(row=17, column=3, value="=C12+C16").number_format = F.FMT_NUMBER

    def render_wacc(self):
        """Formula-driven WACC build.

        Blue cells (C3–C7, C9, C14) are hardcoded inputs read from the engine.
        All other rows are live Excel formulas:
          C8  = after-tax Kd         = C7*(1-C6)
          C10 = debt weight          = 1-C9
          C11 = current levered beta = C5*(1+(1-C6)*(C10/C9))
          C12 = Ke (CAPM, current)   = C3+C11*C4
          C13 = WACC (current)       = C9*C12+C10*C8
          C15 = target levered beta  = C5*(1+(1-C6)*(C14/(1-C14)))
          C16 = Ke (CAPM, target)    = C3+C15*C4
          C17 = WACC (target)        = (1-C14)*C16+C14*C8
          C18 = WACC (avg, used)     = (C13+C17)/2  ← linked by 8_FCF_DCF

        Editing any blue input cell recomputes the full WACC and propagates
        through to Enterprise Value and Intrinsic Value automatically.
        """
        ws = self._create_sheet("9_WACC")
        self._write_header(ws, 2, 2, "WACC Build — CAPM / Hamada")
        w = self.results["wacc"]
        pct = Formats.FMT_PERCENT
        r = 3
        # ── Blue inputs (C3–C9, C14) ─────────────────────────────────────────
        r = self._kv(ws, r, "Risk-free rate (rf)",            w["rf"],                    pct,  blue=True)  # C3
        r = self._kv(ws, r, "Total ERP",                      w["total_erp"],             pct,  blue=True)  # C4
        r = self._kv(ws, r, "Unlevered (asset) beta",         w["median_asset_beta"],    '0.00', blue=True)  # C5
        r = self._kv(ws, r, "Tax rate",                       w["tax_rate"],              pct,  blue=True)  # C6
        r = self._kv(ws, r, "Pre-tax cost of debt",           w["pretax_kd"],             pct,  blue=True)  # C7
        # ── Formula rows ─────────────────────────────────────────────────────
        r = self._kv(ws, r, "After-tax cost of debt",         "=C7*(1-C6)",               pct)             # C8
        r = self._kv(ws, r, "Equity weight (market value)",   w["current_equity_weight"], pct,  blue=True)  # C9
        r = self._kv(ws, r, "Debt weight (book)",             "=1-C9",                    pct)             # C10
        stock_beta = w.get("stock_beta")
        if stock_beta:
            r = self._kv(ws, r, "Current levered beta",       stock_beta,                 '0.00', blue=True)  # C11
        else:
            r = self._kv(ws, r, "Current levered beta",       "=C5*(1+(1-C6)*(C10/C9))", '0.00')          # C11
        r = self._kv(ws, r, "Cost of equity (CAPM)",          "=C3+C11*C4",               pct)             # C12
        r = self._kv(ws, r, "WACC (current structure)",       "=C9*C12+C10*C8",           pct)             # C13
        r = self._kv(ws, r, "Target debt-to-cap",             w["target_debt_to_cap"],    pct,  blue=True)  # C14
        r = self._kv(ws, r, "Target levered beta",            "=C5*(1+(1-C6)*(C14/(1-C14)))", '0.00')      # C15
        r = self._kv(ws, r, "Cost of equity (target)",        "=C3+C15*C4",               pct)             # C16
        r = self._kv(ws, r, "WACC (target structure)",        "=(1-C14)*C16+C14*C8",      pct)             # C17
        r = self._kv(ws, r, "WACC (average, used)",           "=(C13+C17)/2",             pct)             # C18  ← 8_FCF_DCF!$C$9 links here

    def render_terminal(self):
        ws = self._create_sheet("10_Terminal")
        self._write_header(ws, 2, 2, f"Terminal Value ({self._unit})")
        t = self.results["terminal"]
        r = 3
        r = self._kv(ws, r, "Terminal growth (g)", t["terminal_growth"], Formats.FMT_PERCENT, blue=True)
        r = self._kv(ws, r, "Exit EV/EBITDA multiple", t["exit_multiple"], Formats.FMT_MULTIPLE, blue=True)
        r = self._kv(ws, r, "Gordon-growth TV", t["gordon_tv"])
        r = self._kv(ws, r, "Exit-multiple TV", t["multiple_tv"])
        r = self._kv(ws, r, "Average TV (used)", t["avg_tv"])
                
    def render_bridge(self):
        """Live EV→Equity bridge. EV is linked from the FCF/DCF sheet; the bridge
        items are editable; Equity Value and Intrinsic/Share are formulas. Editing
        any driver upstream flows through to the intrinsic value here (C13)."""
        F = Formats
        ws = self._create_sheet("11_Valuation_Bridge")
        c = ws.cell(row=2, column=2, value=f"Enterprise → Equity Bridge ({self._unit})")
        c.font = F.FONT_HEADER; c.fill = F.FILL_HEADER
        b = self.results["bridge"]
        shares = self.results["shares"]["diluted_shares"]

        # (label, value, kind)  rows 3..10 sum to Equity Value
        items = [
            ("Enterprise Value", "='8_FCF_DCF'!$C$17", "link"),
            ("Add: Cash", b["cash"], "in"),
            ("Less: Debt", -b["debt"], "in"),
            ("Less: NCI", -b["nci"], "in"),
            ("Less: Preferred", -b["preferred"], "in"),
            ("Add: Investments", b["investments"], "in"),
            ("Less: Pension", -b["pension"], "in"),
            ("Add: NOLs", b["nols"], "in"),
        ]
        r = 3
        for label, val, kind in items:
            ws.cell(row=r, column=2, value=label).font = F.FONT_BOLD
            cell = ws.cell(row=r, column=3, value=val)
            cell.number_format = F.FMT_NUMBER
            cell.font = F.FONT_LINK if kind == "link" else F.FONT_INPUT
            r += 1
        ws.cell(row=11, column=2, value="Equity Value").font = F.FONT_BOLD
        ws.cell(row=11, column=3, value="=SUM(C3:C10)").number_format = F.FMT_NUMBER
        ws.cell(row=12, column=2, value="Diluted Shares").font = F.FONT_BOLD
        sc = ws.cell(row=12, column=3, value=shares); sc.number_format = '#,##0'; sc.font = F.FONT_INPUT
        ws.cell(row=13, column=2, value=f"Intrinsic Value / Share ({self._cur})").font = Formats.FONT_BOLD
        iv = ws.cell(row=13, column=3, value="=C11*1000000/C12")
        iv.number_format = '#,##0.00'; iv.font = F.FONT_BOLD
        
        # DDM Cross check
        ws.cell(row=15, column=2, value="DDM Cross-Check (dividend-sensitive)").font = F.FONT_BOLD
        ddm = self.results.get("ddm", {})
        if ddm.get("applicable"):
            ddm_val = ddm["ddm_intrinsic_per_share"]
            dc = ws.cell(row=15, column=3, value=ddm_val)
            dc.number_format = '#,##0.00'
            dc.font = F.FONT_BOLD
        else:
            ws.cell(row=15, column=3, value=f"N/A ({ddm.get('reason', '')})")
            
    def render_sensitivity(self):
        ws = self._create_sheet("12_Sensitivity")
        c = ws.cell(row=2, column=2, value="Intrinsic / Share — WACC (rows) x Terminal g (cols)")
        c.font = Formats.FONT_HEADER
        c.fill = Formats.FILL_HEADER
        p, fc, w, t = self.results["proj"], self.results["fcf"], self.results["wacc"], self.results["terminal"]
        ufcf = p["ufcf"]
        n = len(ufcf)
        final_ufcf = ufcf[-1] if ufcf else 0.0
        # Net non-EV adjustments (cash/debt/NCI/…) implied by the bridge, held constant
        bridge_add = self.results["bridge"]["equity_value"] - fc["enterprise_value"]
        shares = self.results["shares"]["diluted_shares"] or 1.0
        base_w, base_g = w["avg_wacc"], t["terminal_growth"]
        waccs = [base_w - 0.02, base_w - 0.01, base_w, base_w + 0.01, base_w + 0.02]
        gs = [base_g - 0.01, base_g - 0.005, base_g, base_g + 0.005, base_g + 0.01]
        ws.cell(row=4, column=2, value="WACC \\ g").font = Formats.FONT_BOLD
        for j, g in enumerate(gs):
            hc = ws.cell(row=4, column=3 + j, value=g)
            hc.number_format = Formats.FMT_PERCENT
            hc.font = Formats.FONT_BOLD
        for i, wc in enumerate(waccs):
            lc = ws.cell(row=5 + i, column=2, value=wc)
            lc.number_format = Formats.FMT_PERCENT
            lc.font = Formats.FONT_BOLD
            for j, g in enumerate(gs):
                pv = sum(cf / ((1 + wc) ** (k + 0.5)) for k, cf in enumerate(ufcf))
                tv = (final_ufcf * (1 + g) / (wc - g)) if wc > g else 0.0
                pv_tv = tv / ((1 + wc) ** (n - 0.5)) if n else 0.0
                eq = pv + pv_tv + bridge_add
                ips = eq * 1e6 / shares
                cell = ws.cell(row=5 + i, column=3 + j, value=ips)
                cell.number_format = Formats.FMT_NUMBER
                if abs(wc - base_w) < 1e-9 and abs(g - base_g) < 1e-9:
                    cell.fill = Formats.FILL_SUBHEADER  # highlight base case

    def render_sources(self):
        ws = self._create_sheet("13_Sources")
        c = ws.cell(row=2, column=2, value="Sources & Assumption Basis")
        c.font = Formats.FONT_HEADER
        c.fill = Formats.FILL_HEADER
        r = 3
        ws.cell(row=r, column=2, value="Documents ingested by Gemini (AJP):").font = Formats.FONT_BOLD
        r += 1
        srcs = list(self.ajp.meta.sources_ingested or [])
        if not srcs:
            srcs = ["(none — engine used historical-data defaults; AJP not supplied)"]
        for s in srcs:
            ws.cell(row=r, column=2, value=str(s))
            r += 1
        r += 1
        ws.cell(row=r, column=2, value="Key assumptions used & basis:").font = Formats.FONT_BOLD
        r += 1
        for k, v in (self.results["proj"].get("assumptions_used") or {}).items():
            ws.cell(row=r, column=2, value=k)
            cell = ws.cell(row=r, column=3, value=v)
            cell.number_format = Formats.FMT_PERCENT if (isinstance(v, float) and abs(v) < 2) else Formats.FMT_NUMBER
            r += 1
        r += 1
        ws.cell(row=r, column=2, value=f"AJP assumptions provided: {len(self.ajp.assumptions)}").font = Formats.FONT_BOLD

def create_workbook(engine_results: Dict[str, Any], ajp: AJP, filepath: str):
    renderer = WorkbookRenderer(engine_results, ajp)
    wb = renderer.render()
    wb.save(filepath)
