from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class Formats:
    # Fonts
    FONT_NORMAL = Font(name='Arial', size=10, color='000000')
    FONT_BOLD = Font(name='Arial', size=10, bold=True, color='000000')
    FONT_INPUT = Font(name='Arial', size=10, color='0000FF') # Blue for hardcoded inputs
    FONT_LINK = Font(name='Arial', size=10, color='008000')  # Green for cross-sheet
    FONT_HEADER = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    
    # Fills
    FILL_HEADER = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    FILL_SUBHEADER = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    FILL_FLAGGED = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
    
    # Alignments
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    
    # Borders
    BORDER_BOTTOM = Border(bottom=Side(style='thin', color='000000'))
    BORDER_TOP_BOTTOM = Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    
    # Number Formats
    FMT_NUMBER = '#,##0'
    FMT_CURRENCY = '#,##0'
    FMT_PERCENT = '0.0%'
    FMT_MULTIPLE = '0.0"x"'
